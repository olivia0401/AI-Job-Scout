# -*- coding: utf-8 -*-
"""AI Job Scout — 扫描全部工签雇主的 AI/LLM 岗位。

用法:  python app.py  然后浏览器打开 http://127.0.0.1:5050
"""
import contextvars
import csv as csvmod
import hashlib
import hmac
import json
from collections import Counter, deque
import os
import re
import secrets
import sys
import threading
import time
import webbrowser
import xml.etree.ElementTree as ET
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
from datetime import date, datetime, timedelta
from io import BytesIO
from urllib.parse import quote_plus

import requests
from flask import (Flask, Response, g, jsonify, redirect, render_template,
                   request, send_file, session, url_for)

FROZEN = getattr(sys, "frozen", False)  # PyInstaller 打包后的 exe 模式
if FROZEN:
    BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))  # data/ 放在 exe 旁边
    RES_DIR = sys._MEIPASS  # 模板等资源被打进包里
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    RES_DIR = BASE_DIR
DATA_DIR = os.path.join(BASE_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)
STATE_FILE = os.path.join(DATA_DIR, "state.json")
COMPANIES_FILE = os.path.join(DATA_DIR, "companies.json")
ATS_CACHE_FILE = os.path.join(DATA_DIR, "ats_cache.json")
SLUG_OVERRIDE_FILE = os.path.join(DATA_DIR, "slug_overrides.json")
API_KEYS_FILE = os.path.join(DATA_DIR, "api_keys.json")
COMPANY_SIZE_FILE = os.path.join(DATA_DIR, "company_sizes.json")

DEFAULT_KEYWORDS = [
    "ai engineer", "machine learning", "ml engineer", "artificial intelligence",
    "llm", "large language model", "nlp", "deep learning", "genai", "gen ai",
    "generative ai", "foundation model", "prompt engineer", "agentic",
    "ai agent", "conversational ai", "data scientist", "research engineer",
    "computer vision", "mlops", "ai scientist", "applied ai",
    # 扩充：召回更多“标题不含 ai/ml 但其实是 AI 方向”的岗位
    "applied scientist", "research scientist", "recommendation",
    "recommendations", "recommender", "personalization", "search ranking",
    "learning to rank", "information retrieval", "knowledge graph",
    "llmops", "ai platform", "model evaluation", "reinforcement learning",
]

# 这些词在标题里是强信号，但在 JD 正文里到处都是（"make recommendations to
# stakeholders"、营销文案里的 personalization），只允许参与标题匹配
TITLE_ONLY_KEYWORDS = {"recommendation", "recommendations", "personalization"}

# 命中即标 ⭐LLM 的关键词（用户重点关注）
LLM_KEYWORDS = [
    "llm", "large language model", "genai", "gen ai", "generative ai",
    "foundation model", "prompt engineer", "agentic", "ai agent", "rag",
    "conversational ai", "chatbot", "nlp",
]

UK_RE = re.compile(r"\b(united kingdom|uk|london|manchester|birmingham|cambridge|oxford|"
                   r"edinburgh|glasgow|bristol|leeds|belfast|cardiff|reading|milton keynes|"
                   r"newcastle|sheffield|liverpool|nottingham|southampton|brighton|"
                   r"england|scotland|wales)\b")

# 欧洲：按对 AI 人才需求排序靠前的国家优先匹配
EU_PLACES = [
    ("Germany", r"germany|berlin|munich|münchen|frankfurt|hamburg|cologne|stuttgart"),
    ("Netherlands", r"netherlands|amsterdam|rotterdam|eindhoven|utrecht|the hague"),
    ("Ireland", r"ireland|dublin|cork"),
    ("France", r"france|paris|lyon|grenoble|toulouse|nice"),
    ("Switzerland", r"switzerland|zurich|zürich|geneva|lausanne|basel"),
    ("Sweden", r"sweden|stockholm|gothenburg"),
    ("Denmark", r"denmark|copenhagen"),
    ("Spain", r"spain|madrid|barcelona|valencia"),
    ("Portugal", r"portugal|lisbon|porto"),
    ("Poland", r"poland|warsaw|krakow|kraków|wroclaw|gdansk"),
    ("Austria", r"austria|vienna"),
    ("Belgium", r"belgium|brussels|antwerp|ghent"),
    ("Finland", r"finland|helsinki"),
    ("Norway", r"norway|oslo"),
    ("Italy", r"italy|milan|rome|turin"),
    ("Czechia", r"czech|prague"),
    ("Luxembourg", r"luxembourg"),
    ("Estonia", r"estonia|tallinn"),
    ("Lithuania", r"lithuania|vilnius"),
    ("Greece", r"greece|athens"),
    ("Hungary", r"hungary|budapest"),
    ("Romania", r"romania|bucharest|cluj"),
    ("Europe", r"europe|emea|european union"),
]
EU_RES = [(label, re.compile(r"\b(" + pat + r")\b")) for label, pat in EU_PLACES]

CHINA_RE = re.compile(
    r"\b(china|shanghai|beijing|shenzhen|hangzhou|guangzhou|nanjing|chengdu|wuhan|"
    r"suzhou|xi'an|xian|tianjin|hong kong|hongkong|macau|taipei|taiwan)\b")

OTHER_RE = re.compile(
    r"\b(usa|us|u\.s\.|united states|canada|toronto|vancouver|montreal|india|bangalore|"
    r"bengaluru|hyderabad|mumbai|delhi|pune|"
    r"japan|tokyo|korea|seoul|singapore|australia|sydney|melbourne|new zealand|brazil|"
    r"sao paulo|mexico|argentina|chile|colombia|qatar|doha|dubai|uae|abu dhabi|saudi|"
    r"israel|tel aviv|turkey|istanbul|egypt|cairo|nigeria|lagos|kenya|nairobi|"
    r"south africa|cape town|johannesburg|pakistan|bangladesh|vietnam|philippines|"
    r"indonesia|thailand|malaysia|new york|san francisco|seattle|austin|boston|chicago|"
    r"denver|los angeles|washington|atlanta|miami|philadelphia|dallas|houston|phoenix)\b")


def classify_region(location):
    """返回 (region, label)：uk / europe / china / remote / other / unknown"""
    loc = (location or "").lower()
    if not loc:
        return ("unknown", "")  # 没写地点 ≠ 英国：单列，前端可选择性显示
    if UK_RE.search(loc):
        return ("uk", "UK")
    for label, rx in EU_RES:
        if rx.search(loc):
            return ("europe", label)
    if CHINA_RE.search(loc):
        return ("china", "China")
    if OTHER_RE.search(loc):
        return ("other", "")
    if "remote" in loc:
        return ("remote", "Remote")
    return ("unknown", "")


LEVEL_LABELS = {
    "grad": "实习/毕业生", "junior": "初级", "mid": "中级",
    "senior": "高级", "staff": "Staff/Lead", "mgmt": "管理层",
}
_LEVEL_RES = [
    ("grad", re.compile(r"\b(intern|internship|graduate|grad|early careers?|entry[- ]?level|apprentice|working student|placement)\b")),
    ("mgmt", re.compile(r"\b(head|director|vp|vice president|chief|manager|management)\b")),
    ("staff", re.compile(r"\b(staff|principal|distinguished|fellow|lead)\b")),
    ("senior", re.compile(r"\b(senior|sr)\b")),
    ("junior", re.compile(r"\b(junior|jr|associate)\b")),
]


def classify_level(title):
    t = (title or "").lower()
    for level, rx in _LEVEL_RES:
        if rx.search(t):
            return level
    return "mid"

HEADERS = {"User-Agent": "Mozilla/5.0 (personal job-search tool)"}
# (连接, 读取) 分离：死掉的每公司子域名 2 秒就放弃连接（原来连接也要磨满 5 秒）；
# 读取仍给 5 秒，保证 Workable 这类偶尔较慢但有效的招聘页不会被误判为“无岗位”。
TIMEOUT = (2, 5)
# 外层并发。全量热路径只跑 4 家健康 ATS(GH/Lever/Ashby/SR)——它们不限流、且探测在工作
# 线程内顺序跑以复用 keep-alive 连接，所以可以放到 32。限流最狠的 Workable 已移出热路径
# （见 WORKABLE_PROBES），不会再被高并发打成 429。
SCAN_WORKERS = 32
# Workable 全局并发上限：命中最多但限流最狠。它是 deep 补扫的吞吐闸门（每家 find_ats 都要先
# 过这道信号量）——4 太保守把 32 线程压成近似单线程。提到 12：3× 吞吐；配合单探测 429 退避
# 自节流（见 probe_workable）+ 已废除整体 5 分钟冷却，实测更快。若频繁 429 可调回小一点。
WORKABLE_CONC = 12
SWEEP_SLUGS = 2          # 全量 sweep 每家公司试几个 slug 候选（命中基本落在前 2 个）
# 深度补扫专用：无限流的 Recruitee/Personio 纯 DNS/IO 等待，可放更高并发把 Workable 排队的
# 空档填满；且补扫求快，每家只试 1 个 slug（省一半探测）。
DEEP_WORKERS = 48
DEEP_SLUGS = 1
LONG_RUNNING_DAYS = 30   # 岗位在线 ≥30 天算“长期在招”
NEW_DAYS = 3             # 首次发现 ≤3 天算“新岗位”

app = Flask(__name__, template_folder=os.path.join(RES_DIR, "templates"))
# 模板改动后浏览器刷新即生效，不必重启服务器（前端调整迭代更省心）
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.jinja_env.auto_reload = True

# 会话密钥：持久化到 data/，重启后已登录用户不会掉线
_SECRET_FILE = os.path.join(DATA_DIR, "secret.key")
if os.path.exists(_SECRET_FILE):
    with open(_SECRET_FILE, encoding="utf-8") as _f:
        app.secret_key = _f.read().strip()
else:
    app.secret_key = secrets.token_hex(32)
    with open(_SECRET_FILE, "w", encoding="utf-8") as _f:
        _f.write(app.secret_key)
app.permanent_session_lifetime = timedelta(days=30)

# ---- 可观测性：RED 指标（Rate / Errors / Duration）--------------------------
# 记录最近请求的耗时与状态码，暴露两个端点：
#   /stats   —— 人可读 JSON：p50/p95/p99、错误率、吞吐、uptime（可截图当证据）
#   /metrics —— Prometheus 文本格式：以后接 Grafana 画图用
_metrics_lock = threading.Lock()
_lat_samples = deque(maxlen=2000)   # 最近 N 次请求耗时（毫秒），环形缓冲省内存
_req_total = 0                      # 累计请求数
_req_by_status = Counter()          # 按 HTTP 状态码计数
_req_errors = 0                     # 5xx 数
_proc_start = time.time()           # 进程启动时刻（算 uptime）
_METRICS_SKIP = ("/stats", "/metrics", "/healthz")  # 这些自身不计入，避免污染真实流量


@app.before_request
def _metrics_start():
    g._t0 = time.time()


@app.after_request
def _metrics_end(resp):
    global _req_total, _req_errors
    t0 = g.get("_t0")
    if t0 is not None and request.path not in _METRICS_SKIP:
        dur_ms = (time.time() - t0) * 1000.0
        with _metrics_lock:
            _lat_samples.append(dur_ms)
            _req_total += 1
            _req_by_status[resp.status_code] += 1
            if resp.status_code >= 500:
                _req_errors += 1
    return resp


def _percentile(sorted_vals, p):
    """从已排序列表取第 p 百分位（p=95 表示 p95）。空列表返回 0。"""
    if not sorted_vals:
        return 0.0
    k = min(len(sorted_vals) - 1, int(round((p / 100.0) * (len(sorted_vals) - 1))))
    return sorted_vals[k]


@app.route("/stats")
def stats():
    with _metrics_lock:
        lat = sorted(_lat_samples)
        total, errors = _req_total, _req_errors
        by_status = dict(_req_by_status)
    uptime = time.time() - _proc_start
    return jsonify({
        "uptime_seconds": round(uptime),
        "requests_total": total,
        "requests_per_min": round(total / (uptime / 60), 2) if uptime > 0 else 0,
        "errors_5xx": errors,
        "error_rate": round(errors / total, 4) if total else 0.0,
        "latency_ms": {
            "p50": round(_percentile(lat, 50), 1),
            "p95": round(_percentile(lat, 95), 1),
            "p99": round(_percentile(lat, 99), 1),
            "max": round(lat[-1], 1) if lat else 0.0,
        },
        "by_status": by_status,
        "samples": len(lat),
    })


@app.route("/metrics")
def metrics():
    with _metrics_lock:
        lat = sorted(_lat_samples)
        total, errors = _req_total, _req_errors
    uptime = time.time() - _proc_start
    lines = [
        "# HELP jobscout_requests_total Total HTTP requests handled.",
        "# TYPE jobscout_requests_total counter",
        f"jobscout_requests_total {total}",
        "# HELP jobscout_request_errors_total HTTP 5xx responses.",
        "# TYPE jobscout_request_errors_total counter",
        f"jobscout_request_errors_total {errors}",
        "# HELP jobscout_uptime_seconds Process uptime in seconds.",
        "# TYPE jobscout_uptime_seconds gauge",
        f"jobscout_uptime_seconds {round(uptime)}",
        "# HELP jobscout_request_latency_ms Request latency percentiles in milliseconds.",
        "# TYPE jobscout_request_latency_ms gauge",
        f'jobscout_request_latency_ms{{quantile="0.5"}} {round(_percentile(lat, 50), 1)}',
        f'jobscout_request_latency_ms{{quantile="0.95"}} {round(_percentile(lat, 95), 1)}',
        f'jobscout_request_latency_ms{{quantile="0.99"}} {round(_percentile(lat, 99), 1)}',
    ]
    return Response("\n".join(lines) + "\n", mimetype="text/plain; version=0.0.4")


# ---- 可选登录保护（部署到公网时用）----------------------------------------
# ---------------------------------------------------------------- 多用户登录 / 注册
# 多用户模式：设了环境变量 APP_MULTIUSER=1，或 HOST=0.0.0.0（对外暴露）时自动开启。
#   → 需要登录；任何人可在 /register 自助注册；密码加盐哈希(pbkdf2)存 data/users.json。
# 本地默认（HOST=127.0.0.1、未设 APP_MULTIUSER）→ 单机免登录，行为跟以前完全一样。
# ⚠️ 开放注册 + 共用你的 API key：务必配合每人每日用量上限（阶段 3），否则会被刷爆。
USERS_FILE = os.path.join(DATA_DIR, "users.json")
LOCAL_USER = "local"   # 本地免登录时的默认用户名
MULTIUSER = bool(os.environ.get("APP_MULTIUSER")) or os.environ.get("HOST") == "0.0.0.0"
# 开放自助注册。默认关：陌生人无法注册（否则会看到你的数据、烧你的 API key）。
# 等每人独立数据(阶段B)+用量上限(阶段C)做完，再设 REGISTRATION_OPEN=1 开放。
REG_OPEN = bool(os.environ.get("REGISTRATION_OPEN"))
_USERNAME_RE = re.compile(r"^[A-Za-z0-9_.\-]{2,32}$")
_OPEN_PATHS = {"/login", "/register", "/logout", "/healthz"}  # 免登录路径
users = {}  # 用户名 -> {"salt","hash","created"}；明文密码从不落盘
_users_lock = threading.Lock()

# ---- 每人每日用量上限（花钱护栏）------------------------------------------
# 开放注册后，所有 LLM 调用都烧 owner 的 API key。用每人每日配额挡住刷爆。
# 都可用环境变量覆盖；owner(APP_USER) 不受限；本地单机模式不限。
OWNER = os.environ.get("APP_USER") or None
LLM_DAILY_CAP = int(os.environ.get("LLM_DAILY_CAP", "100"))
SCAN_DAILY_CAP = int(os.environ.get("SCAN_DAILY_CAP", "10"))
usage = {}   # uid -> {"day": "YYYY-MM-DD", "llm": int, "scan": int}
_usage_lock = threading.Lock()


class UsageLimitError(Exception):
    """超出每日配额。请求处理里由 errorhandler 转成 429。"""


def _usage_today(uid):
    today = date.today().isoformat()
    u = usage.get(uid)
    if u is None or u.get("day") != today:
        u = {"day": today, "llm": 0, "scan": 0}
        usage[uid] = u
    return u


def _charge(kind, cap, n=1):
    """在配额内则计数；超限抛 UsageLimitError。owner/本地不计。"""
    if not MULTIUSER:
        return
    uid = _uid()
    if OWNER and uid == OWNER:
        return
    with _usage_lock:
        u = _usage_today(uid)
        if u[kind] + n > cap:
            raise UsageLimitError(
                f"Daily {kind} limit reached ({cap}/day). Please try again tomorrow.")
        u[kind] += n


@app.errorhandler(UsageLimitError)
def _handle_usage_limit(e):
    return jsonify({"error": str(e)}), 429


def _hash_pw(password, salt=None):
    if salt is None:
        salt = secrets.token_hex(16)
    dk = hashlib.pbkdf2_hmac("sha256", (password or "").encode("utf-8"),
                             bytes.fromhex(salt), 200_000)
    return salt, dk.hex()


def _verify_pw(password, salt, expected_hex):
    try:
        _, got = _hash_pw(password, salt)
    except Exception:
        return False
    return hmac.compare_digest(got, expected_hex)


def _norm_user_record(v):
    """规整成 {'salt','hash','created'}。兼容明文（老 users.json / 环境变量）→ 载入即哈希。"""
    if isinstance(v, dict) and v.get("salt") and v.get("hash"):
        return {"salt": v["salt"], "hash": v["hash"], "created": v.get("created", 0)}
    if isinstance(v, str) and v:   # 明文密码
        salt, h = _hash_pw(v)
        return {"salt": salt, "hash": h, "created": 0}
    return None


def load_users():
    """从 users.json + 老环境变量载入账号表。data/ 已被 gitignore，凭据不会进仓库。"""
    global users
    loaded = {}
    if os.path.exists(USERS_FILE):
        try:
            with open(USERS_FILE, encoding="utf-8") as f:
                raw = json.load(f)
            for k, v in raw.items():
                rec = _norm_user_record(v)
                if k and rec:
                    loaded[str(k)] = rec
        except Exception:
            loaded = {}
    env_u = os.environ.get("APP_USER", "")
    env_p = os.environ.get("APP_PASSWORD", "")
    if env_u and env_p and env_u not in loaded:   # 兼容老的单账号环境变量
        salt, h = _hash_pw(env_p)
        loaded[env_u] = {"salt": salt, "hash": h, "created": 0}
    users = loaded


def save_users():
    with _users_lock:
        payload = json.dumps(users, ensure_ascii=False, indent=2)
    _atomic_write(USERS_FILE, payload)


def create_user(name, password):
    """新建账号，返回 (ok, 错误信息)。用户名校验 + 密码长度 + 去重。"""
    name = (name or "").strip()
    if not _USERNAME_RE.match(name):
        return False, "Username must be 2–32 characters: letters, digits, or _ . -"
    if len((password or "")) < 6:
        return False, "Password must be at least 6 characters"
    with _users_lock:
        if name in users:
            return False, "That username is already taken"
        salt, h = _hash_pw(password)
        users[name] = {"salt": salt, "hash": h, "created": int(time.time())}
    save_users()
    return True, ""


def _bind_user(uid):
    """把本次请求绑定到某用户：加载其数据 + 设置上下文（供 state[...] 路由）。"""
    g.user = uid
    ensure_user(uid)
    _ctx_user.set(uid)


@app.before_request
def _require_login():
    # 每个请求先重置上下文，防止复用的工作线程串到上一个请求的用户
    _ctx_user.set(None)
    if not MULTIUSER:
        _bind_user(LOCAL_USER)   # 本地单机模式，不登录
        return
    p = request.path
    if p in _OPEN_PATHS or p.startswith("/static/"):
        return  # 这些路径不碰 state；上下文保持 None
    user = session.get("user")
    if user and user in users:
        _bind_user(user)
        return
    # 未登录：页面导航跳登录页，前端 fetch 拿到 401（避免把登录 HTML 当 JSON 解析）
    if request.method == "GET" and "text/html" in request.headers.get("Accept", ""):
        return redirect(url_for("login_page", next=p))
    return jsonify({"error": "not logged in"}), 401


@app.route("/login", methods=["GET", "POST"])
def login_page():
    if not MULTIUSER:
        return redirect("/")
    error = ""
    nxt = request.values.get("next") or "/"
    if not nxt.startswith("/"):
        nxt = "/"
    if request.method == "POST":
        name = (request.form.get("username") or "").strip()
        rec = users.get(name)
        if rec and _verify_pw(request.form.get("password") or "",
                              rec["salt"], rec["hash"]):
            session.permanent = True
            session["user"] = name
            return redirect(nxt)
        error = "Incorrect username or password"
    return render_template("login.html", mode="login", error=error,
                           next=nxt, reg_open=REG_OPEN)


@app.route("/register", methods=["GET", "POST"])
def register_page():
    if not MULTIUSER:
        return redirect("/")
    if not REG_OPEN:
        return redirect(url_for("login_page"))  # 注册未开放 → 回登录页
    error = ""
    if request.method == "POST":
        name = (request.form.get("username") or "").strip()
        pw = request.form.get("password") or ""
        if pw != (request.form.get("password2") or ""):
            error = "The two passwords don't match"
        else:
            ok, msg = create_user(name, pw)
            if ok:
                session.permanent = True
                session["user"] = name
                return redirect("/")
            error = msg
    return render_template("login.html", mode="register", error=error,
                           next="/", reg_open=REG_OPEN)


@app.route("/logout")
def logout_page():
    session.pop("user", None)
    return redirect(url_for("login_page"))

# ---------------------------------------------------------------- state（多租户）
# 每个用户一份独立的 state（简历/岗位/匹配/投递…各自独立）。
# `state` 是一个按「当前用户」路由的代理：所有 state[...] 访问自动落到当前用户的数据上。
# 当前用户由 contextvars 决定——请求线程在 before_request 里设，后台线程在入口处设。
# 公司总目录（含 12 万赞助商）是所有用户共享的一份对象，避免重复占内存/硬盘。
_lock = threading.Lock()                 # 保护 state 写入（全局粗粒度锁，够用）
_states_lock = threading.Lock()          # 保护 user_states 字典本身
user_states = {}                         # uid -> 该用户的 state dict
_shared_companies = {}                   # 公司总目录：所有用户共享同一个对象
_ctx_user = contextvars.ContextVar("ctx_user", default=None)


def _uid():
    """当前上下文用户；未设置时回落到 LOCAL_USER（本地单机 / 免登录路径）。"""
    return _ctx_user.get() or LOCAL_USER


def blank_state():
    """一份全新的用户 state。companies 指向共享目录（同一对象，不复制 12 万条）。"""
    return {
        "companies": _shared_companies,   # 共享引用（所有用户同一份公司目录）
        "keywords": list(DEFAULT_KEYWORDS),
        "exclude_keywords": [],  # 标题命中即隐藏（如 sales / frontend / qa）
        "target_titles": [],     # 目标岗位标题（空=用 DEFAULT_TARGET_TITLES）
        "acceptable_titles": [], # 可接受岗位标题（空=用 DEFAULT_ACCEPTABLE_TITLES）
        "uk_only": True,
        "auto_hours": 0,   # >0 时后台每隔 N 小时自动快速刷新
        "results": {},     # name -> result（只存探测到招聘页的公司 + 精选公司）
        "jobs_seen": {},   # job_url -> {"first_seen","last_seen"}
        "scan": {"running": False, "done": 0, "total": 0, "mode": None,
                 "started_ts": 0, "stop": False, "new_jobs": 0},
        "last_auto": 0,
        "resume": None,        # {"text","keywords","updated","name"}
        "match_scores": {},    # job_url -> {"score","hit","must_miss","nice_miss","flags","n"}
        "match": {"running": False, "done": 0, "total": 0},
        "match_stats": {},     # JD 抓取覆盖率：{"jd_ok","jd_fail","updated"}
        "exp_years": 1,        # 我的工作年限（用于年限门槛检测）
        "agg_jobs": {},        # 聚合器(Adzuna/Reed)/社媒招聘帖岗位：url -> {...}
        "agg": {"running": False, "done": 0, "total": 0},
        "social": {"running": False, "done": 0, "total": 0},  # 社媒招聘帖(HN)抓取进度
        "tracker": {},         # 投递清单：job_url -> {"status","ts","title","company","location"}
        "size": {"running": False, "done": 0, "total": 0},  # 公司规模识别进度
        "hidden": {},          # 用户手动隐藏的岗位：url -> {"ts","title"}
        "advice": {},          # 岗位简历建议缓存：url -> markdown（刷新页面不丢、不重复扣 LLM 费）
        "outreach": {},        # networking 草稿缓存：url -> markdown
    }


def ensure_user(uid):
    """确保该用户的 state 已在内存（首次访问时从磁盘懒加载）。返回其 state dict。"""
    with _states_lock:
        st = user_states.get(uid)
        if st is not None:
            return st
    st = blank_state()          # 锁外构建 + 读盘，避免半成品被别的线程看到
    load_user_state(uid, st)
    with _states_lock:
        existing = user_states.get(uid)
        if existing is not None:  # 竞态：别的线程已建好，用它的
            return existing
        user_states[uid] = st
        return st


class _StateProxy:
    """把全局名字 `state` 变成「当前用户的 state」。只需 __getitem__/__setitem__/get。"""
    def _d(self):
        uid = _uid()
        st = user_states.get(uid)
        return st if st is not None else ensure_user(uid)

    def __getitem__(self, k):
        return self._d()[k]

    def __setitem__(self, k, v):
        self._d()[k] = v

    def get(self, k, default=None):
        return self._d().get(k, default)


state = _StateProxy()


def _spawn(fn, *args):
    """启动一个后台线程，并把「当前用户」绑定进新线程的上下文。
    新线程默认不继承 contextvars，所以这里显式捕获 + 设置，避免跑到别人的数据上。"""
    u = _uid()

    def runner():
        _ctx_user.set(u)
        fn(*args)

    threading.Thread(target=runner, daemon=True).start()


# name -> {"ats","slug"}：找到招聘系统
#       | "none"      ：普通探测没找到（深度补扫还会再试它）
#       | "none_deep" ：深度补扫也探过、确认没有（深度补扫不再重复扫 → 停了能续跑）
ats_cache = {}
_NO_ATS = ("none", "none_deep")   # 两者都表示"没有招聘页"
company_sizes = {}  # company name -> {"band": startup|mid|large|unknown, "note": str}
api_keys = {}   # {"adzuna_app_id","adzuna_app_key","reed_api_key"}
slug_overrides = {}  # name.lower() -> {"ats","slug"}：手动修正 slug 猜错的重点公司


def _valid_override(v):
    """一条手动修正是否可用。Workday 用 host+site 定位，其它 ATS 用 slug。"""
    if not v.get("ats"):
        return False
    if v["ats"] == "Workday":
        return bool(v.get("host") and v.get("site"))
    return bool(v.get("slug"))


def _user_dir(uid):
    d = os.path.join(DATA_DIR, "users", uid)
    os.makedirs(d, exist_ok=True)
    return d


def _user_state_file(uid):
    return os.path.join(_user_dir(uid), "state.json")


def load_all():
    """启动时载入所有用户共享的数据：探测缓存 + 公司总目录。各用户 state 懒加载。"""
    global ats_cache, slug_overrides, api_keys, company_sizes
    if os.path.exists(COMPANY_SIZE_FILE):
        try:
            with open(COMPANY_SIZE_FILE, encoding="utf-8") as f:
                company_sizes = json.load(f)
        except Exception:
            company_sizes = {}
    if os.path.exists(API_KEYS_FILE):
        try:
            with open(API_KEYS_FILE, encoding="utf-8") as f:
                api_keys = json.load(f)
        except Exception:
            api_keys = {}
    if os.path.exists(ATS_CACHE_FILE):
        try:
            with open(ATS_CACHE_FILE, encoding="utf-8") as f:
                ats_cache = json.load(f)
        except Exception:
            ats_cache = {}
    if os.path.exists(SLUG_OVERRIDE_FILE):
        try:
            with open(SLUG_OVERRIDE_FILE, encoding="utf-8") as f:
                raw = json.load(f)
            slug_overrides = {k.strip().lower(): v for k, v in raw.items()
                              if isinstance(v, dict) and _valid_override(v)}
        except Exception:
            slug_overrides = {}
    # 公司总目录（共享）：就地填充 _shared_companies，保持 blank_state 里的引用有效
    if os.path.exists(COMPANIES_FILE):
        try:
            with open(COMPANIES_FILE, encoding="utf-8") as f:
                _shared_companies.update(json.load(f))
        except Exception:
            pass
    migrate_legacy_state()


def migrate_legacy_state():
    """把老的单租户 data/state.json 迁移成 owner 账号的 per-user state（只做一次）。
    老文件里的 companies 并入共享目录。原文件保留作备份，不删除。"""
    if not os.path.exists(STATE_FILE):
        return
    owner = os.environ.get("APP_USER") or LOCAL_USER
    dest = _user_state_file(owner)
    if os.path.exists(dest):
        return  # 已迁移
    try:
        with open(STATE_FILE, encoding="utf-8") as f:
            saved = json.load(f)
    except Exception:
        return
    comps = saved.pop("companies", {})
    if isinstance(comps, list):  # 更老的格式：当时导入的就是精选清单
        comps = {c["name"]: {"tags": ["curated"], "town": ""} for c in comps}
    for k, v in (comps or {}).items():
        _shared_companies.setdefault(k, v)
    _atomic_write(dest, json.dumps(saved, ensure_ascii=False))


def load_user_state(uid, st):
    """把某用户的 per-user state.json 载入到 st（companies 不在此文件里，是共享的）。"""
    path = _user_state_file(uid)
    if not os.path.exists(path):
        return
    try:
        with open(path, encoding="utf-8") as f:
            saved = json.load(f)
    except Exception:
        return
    st["keywords"] = saved.get("keywords", list(DEFAULT_KEYWORDS))
    st["exclude_keywords"] = saved.get("exclude_keywords", [])
    st["target_titles"] = saved.get("target_titles", [])
    st["acceptable_titles"] = saved.get("acceptable_titles", [])
    st["uk_only"] = saved.get("uk_only", True)
    st["auto_hours"] = saved.get("auto_hours", 0)
    st["jobs_seen"] = saved.get("jobs_seen", {})
    st["resume"] = saved.get("resume")
    st["match_scores"] = saved.get("match_scores", {})
    st["match_stats"] = saved.get("match_stats", {})
    st["agg_jobs"] = saved.get("agg_jobs", {})
    st["exp_years"] = saved.get("exp_years", 1)
    st["tracker"] = saved.get("tracker", {})
    st["hidden"] = saved.get("hidden", {})
    st["advice"] = saved.get("advice", {})
    st["outreach"] = saved.get("outreach", {})
    for name, r in saved.get("results", {}).items():
        r.pop("links", None)  # 旧版把链接存了盘，现在改为按需生成
        st["results"][name] = r


def _atomic_write(path, payload):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        f.write(payload)
    os.replace(tmp, path)


def save_companies():
    """公司名单单独存（12 万条只在导入时变），避免扫描期间反复写大文件。
    调用方不得持锁。"""
    with _lock:
        payload = json.dumps(state["companies"], ensure_ascii=False)
    _atomic_write(COMPANIES_FILE, payload)


def save_state():
    """把「当前用户」的 state 落盘到 data/users/<uid>/state.json。
    锁内序列化（results/jobs_seen 都不大），锁外写盘。调用方不得持锁。
    companies 是共享的，单独由 save_companies() 存，不进本文件。"""
    dest = _user_state_file(_uid())
    with _lock:
        payload = json.dumps({
            "keywords": state["keywords"],
            "exclude_keywords": state["exclude_keywords"],
            "target_titles": state["target_titles"],
            "acceptable_titles": state["acceptable_titles"],
            "uk_only": state["uk_only"],
            "auto_hours": state["auto_hours"],
            "results": state["results"],
            "jobs_seen": state["jobs_seen"],
            "resume": state["resume"],
            "match_scores": state["match_scores"],
            "match_stats": state["match_stats"],
            "agg_jobs": state["agg_jobs"],
            "exp_years": state["exp_years"],
            "tracker": state["tracker"],
            "hidden": state["hidden"],
            "advice": state["advice"],
            "outreach": state["outreach"],
        }, ensure_ascii=False)
    _atomic_write(dest, payload)


def save_ats_cache():
    with open(ATS_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(ats_cache, f, ensure_ascii=False)


# ---------------------------------------------------------------- slug 生成
STRIP_SUFFIXES = [
    "ltd", "limited", "plc", "llp", "inc", "uk", "group", "holdings",
    "international", "technologies", "technology", "labs", "lab", "co",
]


def slug_candidates(name):
    base = name.lower().strip()
    base = re.sub(r"t/a.*$", " ", base)  # 去掉 trading-as 别名
    base = re.sub(r"[&+]", " and ", base)
    base = re.sub(r"[^a-z0-9\s-]", "", base)
    words = [w for w in base.split() if w]
    trimmed = list(words)
    while len(trimmed) > 1 and trimmed[-1] in STRIP_SUFFIXES:
        trimmed.pop()
    # 只试命中率最高的形式，控制全量清查的请求量
    cands = []
    if trimmed:
        cands.append("".join(trimmed))
        cands.append("-".join(trimmed))
    if words != trimmed and words:
        cands.append("".join(words))
    seen, out = set(), []
    for c in cands:
        if c and c not in seen:
            seen.add(c)
            out.append(c)
    return out[:3]


# ---------------------------------------------------------------- ATS 探测
_tls = threading.local()


def _session():
    if not hasattr(_tls, "s"):
        _tls.s = requests.Session()
        _tls.s.headers.update(HEADERS)
    return _tls.s


class RateLimited(Exception):
    """对方限流/暂时不可用；此时不能把公司记为 none，否则永久漏掉。"""


# 停止扫描的全局开关。深度补扫时在飞任务可能卡在 Workable 信号量排队/退避 sleep 里
# 几十秒到几分钟，光靠 run_scan 主循环检查 state["scan"]["stop"] 根本停不下来——
# 所有慢环节（探测间隙、信号量等待、退避）都要能被这个事件秒级打断。
_stop_event = threading.Event()


# 近期限流(429/503)次数监控——仅作可观测性用（_recent_rl 可供诊断/后续接指标）。
# 注：全源提速后已废除「命中 cap 就整体冷却 5 分钟」的逻辑，429 改由单探测退避自节流。
_rl_lock = threading.Lock()
_rl_hits = []          # 最近的 429/503 时间戳
RL_WINDOW = 120        # 统计窗口（秒）


def _note_rate_limit():
    now = time.time()
    with _rl_lock:
        _rl_hits.append(now)
        cutoff = now - RL_WINDOW
        while _rl_hits and _rl_hits[0] < cutoff:
            _rl_hits.pop(0)


def _recent_rl():
    now = time.time()
    with _rl_lock:
        return sum(1 for t in _rl_hits if t >= now - RL_WINDOW)


def _get_json(url):
    r = _session().get(url, timeout=TIMEOUT)
    if r.status_code in (429, 503):
        _note_rate_limit()
        raise RateLimited(url)
    if r.status_code != 200:
        return None
    return r.json()


def probe_greenhouse(slug):
    data = _get_json(f"https://boards-api.greenhouse.io/v1/boards/{slug}/jobs")
    if not data or "jobs" not in data:
        return None
    jobs = [{
        "title": j.get("title", ""),
        "url": j.get("absolute_url", ""),
        "location": (j.get("location") or {}).get("name", ""),
    } for j in data["jobs"]]
    return {"board_url": f"https://boards.greenhouse.io/{slug}", "jobs": jobs}


def greenhouse_bodies(slug):
    """Greenhouse 基础列表不含 JD，用 content=true 一次拿全公司所有岗位正文。
    只在有“广撒网候选”时才调用，均摊到每家公司最多一次额外请求。"""
    data = _get_json(
        f"https://boards-api.greenhouse.io/v1/boards/{slug}/jobs?content=true")
    out = {}
    for j in (data or {}).get("jobs", []):
        u = j.get("absolute_url", "")
        if u:
            out[u] = _strip_html(j.get("content", ""))
    return out


def _lever_desc(j):
    parts = [j.get("descriptionPlain") or _strip_html(j.get("description", ""))]
    for lst in j.get("lists", []):
        parts.append(lst.get("text", "") + " " + _strip_html(lst.get("content", "")))
    return " ".join(p for p in parts if p)


def probe_lever(slug):
    data = _get_json(f"https://api.lever.co/v0/postings/{slug}?mode=json")
    if not isinstance(data, list):
        return None
    jobs = [{
        "title": j.get("text", ""),
        "url": j.get("hostedUrl", ""),
        "location": (j.get("categories") or {}).get("location", "") or "",
        "desc": _lever_desc(j),  # Lever 列表接口自带 JD，二次确认零成本
    } for j in data]
    return {"board_url": f"https://jobs.lever.co/{slug}", "jobs": jobs}


def probe_ashby(slug):
    data = _get_json(
        f"https://api.ashbyhq.com/posting-api/job-board/{slug}?includeCompensation=false")
    if not data or "jobs" not in data:
        return None
    jobs = [{
        "title": j.get("title", ""),
        "url": j.get("jobUrl", "") or j.get("applyUrl", ""),
        "location": j.get("location", "") or "",
        "desc": j.get("descriptionPlain") or _strip_html(j.get("descriptionHtml", "")),
    } for j in data["jobs"]]
    return {"board_url": f"https://jobs.ashbyhq.com/{slug}", "jobs": jobs}


_WORKABLE_SEM = threading.BoundedSemaphore(WORKABLE_CONC)


def probe_workable(slug):
    # Workable 限流最狠：① 全局并发闸限死同时探测数；② 被 429 退避重试而非直接放弃，
    # 让深度补扫最终能查完（退避期间占着并发槽 → 等于自动降速自愈）。只影响 deep 补扫，
    # 不进全量热路径，所以第一次全量不受它拖累。
    url = f"https://apply.workable.com/api/v1/widget/accounts/{slug}"
    # 排队等并发槽时每 0.5s 看一眼停止开关；停止按 RateLimited 处理 → 不落 none 缓存
    while not _WORKABLE_SEM.acquire(timeout=0.5):
        if _stop_event.is_set():
            raise RateLimited(url)
    try:
        data = None
        for attempt in range(4):
            try:
                data = _get_json(url)
                break
            except RateLimited:
                if attempt == 3:
                    raise                       # 连试 4 次仍限流 → 本轮不落 none，下轮再补
                # 退避：1.5s, 3s, 6s；期间被停止则立刻放弃（wait 返回 True = 事件已置位）
                if _stop_event.wait(1.5 * (2 ** attempt)):
                    raise RateLimited(url)
    finally:
        _WORKABLE_SEM.release()
    if not data or "jobs" not in data:
        return None
    jobs = [{
        "title": j.get("title", ""),
        "url": j.get("url", ""),
        "location": ", ".join(x for x in [j.get("city", ""), j.get("country", "")] if x),
    } for j in data["jobs"]]
    return {"board_url": f"https://apply.workable.com/{slug}/", "jobs": jobs}


def probe_smartrecruiters(slug):
    data = _get_json(f"https://api.smartrecruiters.com/v1/companies/{slug}/postings?limit=100")
    # SmartRecruiters 对不存在的公司也返回 200+空列表，必须把空结果当未命中
    if not data or not data.get("content"):
        return None
    jobs = []
    for j in data["content"]:
        loc = j.get("location") or {}
        jobs.append({
            "title": j.get("name", ""),
            "url": f"https://jobs.smartrecruiters.com/{slug}/{j.get('id', '')}",
            "location": ", ".join(x for x in [loc.get("city", ""), loc.get("country", "")] if x),
        })
    return {"board_url": f"https://jobs.smartrecruiters.com/{slug}", "jobs": jobs}


def probe_recruitee(slug):
    data = _get_json(f"https://{slug}.recruitee.com/api/offers/")
    if not data or "offers" not in data:
        return None
    jobs = [{
        "title": j.get("title", ""),
        "url": j.get("careers_url", ""),
        "location": j.get("location", "") or "",
        "desc": _strip_html(str(j.get("description", "")) + " "
                            + str(j.get("requirements", ""))),
    } for j in data["offers"]]
    return {"board_url": f"https://{slug}.recruitee.com/", "jobs": jobs}


def probe_personio(slug):
    # Personio 每家公司自带公开 XML 岗位源，正文也在里面（二次确认零成本）
    try:
        r = _session().get(f"https://{slug}.jobs.personio.com/xml", timeout=TIMEOUT)
    except Exception:
        return None
    if r.status_code != 200 or "<position" not in r.text:
        return None
    try:
        root = ET.fromstring(r.content)
    except Exception:
        return None
    jobs = []
    for pos in root.iter("position"):
        pid = (pos.findtext("id") or "").strip()
        title = (pos.findtext("name") or "").strip()
        if not pid or not title:
            continue
        office = (pos.findtext("office") or "").strip()
        descs = [jd.findtext("value") or "" for jd in pos.iter("jobDescription")]
        jobs.append({
            "title": title,
            "url": f"https://{slug}.jobs.personio.com/job/{pid}",
            "location": office,
            "desc": _strip_html(" ".join(descs)),
        })
    return {"board_url": f"https://{slug}.jobs.personio.com/", "jobs": jobs}


def probe_workday_tenant(host, site):
    """Workday 公开岗位站的匿名接口。host 如 nvidia.wd5.myworkdayjobs.com，
    site 如 NVIDIAExternalCareerSite。只能通过手动 override 配置（无法从公司名猜）。"""
    if not host or not site:
        return None
    tenant = host.split(".")[0]
    base = f"https://{host}/wday/cxs/{tenant}/{site}"
    jobs, offset = [], 0
    try:
        while offset < 200:  # 最多取 200 个，避免超大站拖慢
            r = _session().post(f"{base}/jobs",
                                json={"appliedFacets": {}, "limit": 20,
                                      "offset": offset, "searchText": ""},
                                timeout=10)
            if r.status_code != 200:
                break
            d = r.json()
            batch = d.get("jobPostings") or []
            for j in batch:
                jobs.append({
                    "title": j.get("title", ""),
                    "url": f"https://{host}/en-US/{site}{j.get('externalPath', '')}",
                    "location": j.get("locationsText", "") or "",
                })  # Workday 列表不含正文 → 只按标题判断
            total = d.get("total") or 0
            offset += 20
            if not batch or offset >= total:
                break
    except Exception:
        return None
    if not jobs:
        return None
    return {"board_url": f"https://{host}/en-US/{site}", "jobs": jobs}


# 健康快探测：走各家共享域名，快且不限流——全量 sweep 默认只跑这几个，能真正跑完
FAST_PROBES = [
    ("Greenhouse", probe_greenhouse),
    ("Lever", probe_lever),
    ("Ashby", probe_ashby),
    ("SmartRecruiters", probe_smartrecruiters),
]
# Workable：命中最多(273家)但限流最狠，放进全量热路径会整片 429 把 sweep 拖到跑不完。
# 移到「深度补扫」(mode=deep) 里，用 _WORKABLE_SEM 节流跑，不拖垮主扫描。
WORKABLE_PROBES = [("Workable", probe_workable)]
# 慢探测：Recruitee/Personio 用「每公司独立子域名」，对不存在的公司要付一次注定
# 失败的 DNS+握手（实测各 ~3.7s），而命中极少（Recruitee 8、Personio 0）。同样放深度补扫。
EXTRA_PROBES = [
    ("Recruitee", probe_recruitee),
    ("Personio", probe_personio),
]
PROBES = FAST_PROBES                             # 默认 hot path（pending/all/known）
CORE_PROBES = FAST_PROBES + WORKABLE_PROBES      # 精选小清单：连 Workable 一起扫（量小不怕限流）
DEEP_PROBES = WORKABLE_PROBES + EXTRA_PROBES     # 深度补扫：Workable + Recruitee + Personio
ALL_PROBES = CORE_PROBES + EXTRA_PROBES
PROBE_FN = dict(ALL_PROBES)       # 缓存命中/手动 override 反查仍需认得全部 ATS
# Workday 不参与按公司名猜 slug（需要 host+site），只能手动 override
OVERRIDE_ATS = set(PROBE_FN) | {"Workday"}

# 预置大厂 Workday 招聘站（host, site）。这些公司用 Workday/自建系统，按公司名猜
# slug 猜不到，故内置 host/site 直连其公开岗位接口。属"尽力而为"：
#   · 探测失败不会有副作用——find_ats 会自动回落到普通探测；
#   · host/site 日后若变动，可在网页「手动修正」里改（会写入 data/slug_overrides.json，覆盖内置）。
# 注意：Google / Amazon / Meta / Microsoft / 摩根等用**自建**招聘系统（非 Workday），
# 无公开接口可抓，仍需用公司卡片里的「Google 直达 / LinkedIn 搜索」手动查。
BUILTIN_WORKDAY = {
    "NVIDIA": ("nvidia.wd5.myworkdayjobs.com", "NVIDIAExternalCareerSite"),
    "Salesforce": ("salesforce.wd12.myworkdayjobs.com", "External_Career_Site"),
    "Adobe": ("adobe.wd5.myworkdayjobs.com", "external_experienced"),
    "Dell Technologies": ("dell.wd1.myworkdayjobs.com", "External"),
    "Mastercard": ("mastercard.wd1.myworkdayjobs.com", "CorporateCareers"),
    "Capital One": ("capitalone.wd1.myworkdayjobs.com", "Capital_One"),
    "Autodesk": ("autodesk.wd1.myworkdayjobs.com", "Ext"),
    "AstraZeneca": ("astrazeneca.wd3.myworkdayjobs.com", "Careers"),
    "Sanofi": ("sanofi.wd3.myworkdayjobs.com", "SanofiCareers"),
    "Workday": ("workday.wd5.myworkdayjobs.com", "Workday"),
}
BUILTIN_OVERRIDES = {name.lower(): {"ats": "Workday", "host": host, "site": site}
                     for name, (host, site) in BUILTIN_WORKDAY.items()}

# 预置一批知名 AI/ML/金融科技雇主，作为 curated 公司种子。它们多用 Greenhouse/Lever/
# Ashby 等可自动探测的系统，加入清单后会被自动识别 slug；探测不到的会显示为"无招聘页"，
# 提醒你手动去查。（这些都在英国招人且常担保工签。）
CURATED_SEED = [
    "Anthropic", "OpenAI", "Cohere", "Hugging Face", "DeepMind", "Databricks",
    "Palantir Technologies", "Wayve", "Faculty", "Quantexa", "Synthesia",
    "Stability AI", "ElevenLabs", "PolyAI", "Speechmatics", "Graphcore",
    "Monzo", "Wise", "Revolut", "Improbable",
] + list(BUILTIN_WORKDAY)


def seed_builtin_companies():
    """把预置大厂/AI 雇主加入公司清单（curated），使其被扫描；Workday 大厂还会命中内置 override。"""
    changed = False
    with _lock:
        existing = {n.strip().lower() for n in state["companies"]}
        for name in CURATED_SEED:
            if name.strip().lower() not in existing:
                state["companies"][name] = {"tags": ["curated"], "town": ""}
                existing.add(name.strip().lower())
                changed = True
    if changed:
        save_companies()


def save_slug_overrides():
    _atomic_write(SLUG_OVERRIDE_FILE, json.dumps(slug_overrides, ensure_ascii=False))


def _probe_override(ov):
    ats = ov.get("ats")
    if ats == "Workday":
        return probe_workday_tenant(ov.get("host", ""), ov.get("site", ""))
    fn = PROBE_FN.get(ats)
    if fn and ov.get("slug"):
        try:
            return fn(ov["slug"])
        except Exception:
            return None
    return None


def find_ats(name, probes=PROBES, max_slugs=SWEEP_SLUGS, force=False):
    """返回 (ats_name, slug, {board_url, jobs}) 或 None。结果写入 ats_cache。
    probes/max_slugs 由扫描模式决定；force=True 时忽略 'none' 缓存（深度补扫用）。
    探测在调用线程内顺序跑——复用该线程的 keep-alive 连接（比每家公司新建线程池、
    每个请求都重握手 TLS 快数倍）。"""
    key = (name or "").strip().lower()
    ov = slug_overrides.get(key) or BUILTIN_OVERRIDES.get(key)  # 用户修正优先，其次内置大厂
    if ov:
        res = _probe_override(ov)
        if res:  # 手动修正优先，抓不到再走自动探测
            return (ov["ats"], ov.get("slug") or ov.get("site") or "", res)
    cached = ats_cache.get(name)
    if cached in _NO_ATS and not force:
        return None
    if isinstance(cached, dict):
        try:
            res = PROBE_FN[cached["ats"]](cached["slug"])
        except Exception:
            res = None
        if res:
            return (cached["ats"], cached["slug"], res)
        # 缓存失效则重新探测

    fallback = None  # 0 岗位的空账号可能是重名壳账号，只作兜底
    rate_limited = False  # 被限流的探测视为"未知"而非"没有"
    for slug in slug_candidates(name)[:max_slugs]:
        for ats_name, fn in probes:
            if _stop_event.is_set():
                rate_limited = True  # 中途放弃按"未知"处理，不写 none 缓存
                break
            try:
                res = fn(slug)
            except RateLimited:
                rate_limited = True
                res = None
            except Exception:
                res = None
            if res:
                if res["jobs"]:
                    ats_cache[name] = {"ats": ats_name, "slug": slug}
                    return (ats_name, slug, res)
                if fallback is None:
                    fallback = (ats_name, slug, res)
        if _stop_event.is_set():
            break
    if fallback:
        ats_cache[name] = {"ats": fallback[0], "slug": fallback[1]}
        return fallback
    if not rate_limited:  # 限流/中途停止时不写缓存，下次续跑会重新探测这家
        # 深度补扫(force=True)也没找到 → 记 none_deep，下次深度补扫直接跳过它。
        # 这样「扫一半停掉、明天接着跑」不会从头再来（普通探测仍记 none，留给深度补扫再试）。
        ats_cache[name] = "none_deep" if force else "none"
    return None


# ---------------------------------------------------------------- 匹配
def kw_match(text, kws):
    return any(re.search(r"\b" + re.escape(k) + r"\b", text) for k in kws)


def kw_hit_count(text, kws):
    """关键词命中总次数（同一个词出现多次都计入）。"""
    return sum(len(re.findall(r"\b" + re.escape(k) + r"\b", text)) for k in kws)


# 广撒网标题：工程/数据/研究类岗位，标题没写 AI 也值得读 JD 正文二次确认
BROAD_TITLE_RE = re.compile(
    r"\b(engineer|developer|scientist|programmer|architect|researcher)\b")

# 标题黑名单：明显不是 AI 研发方向的角色，不给走 JD 正文宽通道——
# 这些岗位挂在"AI 公司"名下时，JD 里的公司简介总会提到 AI，全是误命中
EXCLUDE_TITLE_RE = re.compile(
    r"\b(sales|pre-?sales|account (?:manager|executive)|business development|"
    r"marketing|recruit(?:er|ing|ment)|talent acquisition|customer success|"
    r"support|solutions? (?:architect|engineer|consultant)|"
    r"qa|sdet|quality assurance|test(?:er|ing)?|"
    r"android|ios|mobile|front[- ]?end|"
    r"network|electrical|mechanical|civil|hardware|firmware|embedded|bmc|"
    r"devops|site reliability|sre|security|help ?desk|"
    r"field operations?|control engineer|validation engineer)\b")


def match_job(title, desc, keywords):
    """判断一个岗位是否算 AI 岗，返回命中来源：
      'title' —— 标题直接命中 AI 关键词；
      'jd'    —— 标题是工程/数据/研究类（且不在黑名单里）、JD 正文命中
                 AI 关键词 ≥2 次（拓宽召回；要求 ≥2 次是为了滤掉
                 公司简介里"we are an AI-powered..."这种一嘴带过的套话）；
      None    —— 都没命中。
    JD 正文抓不到时（desc 为空）只走标题，天然保守。"""
    t = (title or "").lower()
    if kw_match(t, keywords):
        return "title"
    if desc and BROAD_TITLE_RE.search(t) and not EXCLUDE_TITLE_RE.search(t):
        body_kws = [k for k in keywords if k not in TITLE_ONLY_KEYWORDS]
        if kw_hit_count(desc.lower(), body_kws) >= 2:
            return "jd"
    return None


# 目标岗位画像：区分"这个岗位方向是否对口"，与"简历技能是否重合"解耦。
# 关键：技能重合高 ≠ 岗位对口——IT Support 的 JD 也会写 python/docker/aws。
DEFAULT_TARGET_TITLES = [
    "machine learning engineer", "ml engineer", "mle",
    "ai engineer", "artificial intelligence engineer",
    "llm engineer", "nlp engineer", "genai engineer", "generative ai engineer",
    "deep learning engineer", "computer vision engineer",
    "machine learning scientist", "ai scientist",
]
DEFAULT_ACCEPTABLE_TITLES = [
    "applied scientist", "research engineer", "research scientist",
    "data scientist", "data science", "applied ai",
    "machine learning researcher", "ml researcher",
    "mlops engineer", "ai platform engineer",
]


def _title_hit(title_l, phrases):
    return any(re.search(r"\b" + re.escape(p) + r"\b", title_l) for p in phrases)


# 词元启发式：AI/ML 方向词 + 角色词。用于兜住写法千变万化的真实岗位标题
# （如 "Software Engineer, Machine Learning" / "ML Platform Engineer" /
#  "NLP Research Engineer" / "Research Scientist, Foundation Models"），
# 避免"精确短语命中"把真实相关岗位误判成 off_target。
_AI_TOKEN_RE = re.compile(
    r"\b(machine learning|ml|deep learning|artificial intelligence|ai|"
    r"llm|large language model|nlp|natural language|computer vision|"
    r"genai|generative ai|foundation model|reinforcement learning|"
    r"recommendations?|recommender|personali[sz]ations?|search ranking|"
    r"mlops|applied ai|data science)\b")
_ENG_ROLE_RE = re.compile(r"\b(engineer|developer|programmer)\b")
_SCI_ROLE_RE = re.compile(r"\b(scientist|researcher|research)\b")


def title_match(title, targets=None, acceptables=None):
    """岗位方向：target=正好想要 / acceptable=能接受 / off_target=方向不对。
    off_target 的岗位无论技能分多高都不该被强推（例如 IT Support、Sales）。

    判定顺序：① 明确黑名单角色(support/sales/qa…)直接 off_target；
    ② 用户显式列的目标/可接受短语优先；③ 词元启发式：AI方向词+角色词——
    兼容各种标题写法，避免误伤真实相关岗位。"""
    t = (title or "").lower()
    tg = targets if targets is not None else DEFAULT_TARGET_TITLES
    ac = acceptables if acceptables is not None else DEFAULT_ACCEPTABLE_TITLES
    if EXCLUDE_TITLE_RE.search(t):          # ① 黑名单：即便带 engineer/AI 也算方向不对
        return "off_target"
    if _title_hit(t, tg):                   # ② 用户/默认目标短语
        return "target"
    if _title_hit(t, ac):
        return "acceptable"
    has_ai = bool(_AI_TOKEN_RE.search(t))   # ③ 词元启发式
    if has_ai and _ENG_ROLE_RE.search(t):
        return "target"
    if has_ai and _SCI_ROLE_RE.search(t):
        return "acceptable"
    return "off_target"


def build_links(name):
    q = quote_plus(name)
    return {
        "li_people": ("https://www.linkedin.com/search/results/people/?keywords="
                      + quote_plus(f'"{name}" (recruiter OR "talent acquisition" OR "hiring manager" OR "engineering manager")')),
        "li_jobs_kw": ("https://www.linkedin.com/jobs/search/?keywords="
                       + quote_plus(f"{name} AI machine learning engineer") + "&location=United%20Kingdom"),
        "google": ("https://www.google.com/search?q="
                   + quote_plus(f'"{name}" ("AI engineer" OR "machine learning engineer") job UK')),
        "hunter": f"https://hunter.io/search/{q}",
        "careers_google": ("https://www.google.com/search?q="
                           + quote_plus(f"{name} careers jobs")),
    }


# ---------------------------------------------------------------- 扫描
def scan_company(name, keywords, uk_only, probes=PROBES, max_slugs=SWEEP_SLUGS,
                 force=False):
    result = {
        "name": name,
        "scanned_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "ats": None, "slug": None, "board_url": None,
        "total_jobs": 0, "matched": [],
        "status": "no_board",
    }
    hit = find_ats(name, probes, max_slugs, force)
    if hit:
        ats_name, slug, res = hit
        result["ats"] = ats_name
        result["slug"] = slug
        result["board_url"] = res["board_url"]
        result["total_jobs"] = len(res["jobs"])
        matched = []
        gh_bodies = None  # Greenhouse 正文按需拉取一次
        for j in res["jobs"]:
            title = (j.get("title") or "").lower()
            body = j.get("desc") or ""
            via = match_job(title, body, keywords)
            if via is None and ats_name == "Greenhouse" and BROAD_TITLE_RE.search(title):
                if gh_bodies is None:
                    gh_bodies = greenhouse_bodies(slug)
                body = gh_bodies.get(j.get("url", ""), "")
                via = match_job(title, body, keywords)
            if via is None:
                continue
            region, rlabel = classify_region(j.get("location"))
            if uk_only and region == "other":
                continue  # 排除美国/亚洲等；英国+欧洲+Remote 都保留
            j.pop("desc", None)  # JD 正文不入库，简历匹配时另存 desc_cache
            j["region"] = region
            j["country"] = rlabel
            j["level"] = classify_level(title)
            j["llm"] = kw_match(title, LLM_KEYWORDS) or bool(
                body and kw_match(body.lower(), LLM_KEYWORDS))
            j["via"] = via  # title=标题命中 / jd=正文命中（标题没写 AI）
            matched.append(j)
        result["matched"] = matched[:80]
        if matched:
            result["status"] = "has_ai"
        elif res["jobs"]:
            result["status"] = "board_no_ai"
        else:
            result["status"] = "no_board"  # 空壳账号
    return result


def _curated_names():
    return {n for n, c in state["companies"].items() if "curated" in c.get("tags", [])}


def _record_result(name, res, today, curated):
    """在锁内调用：登记岗位首见/末见日期，按需入库。"""
    for j in res.get("matched", []):
        seen = state["jobs_seen"].setdefault(j["url"], {"first_seen": today})
        seen["last_seen"] = today
        j["first_seen"] = seen["first_seen"]
        if seen["first_seen"] == today:
            state["scan"]["new_jobs"] += 1
    if res["status"] == "no_board" and name not in curated:
        state["results"].pop(name, None)  # register 里没招聘页的不占地方
    else:
        state["results"][name] = res


def run_scan(names, mode):
    keywords = [k.lower() for k in state["keywords"]]
    uk_only = state["uk_only"]
    curated = _curated_names()
    today = date.today().isoformat()
    # 深度补扫：对已判 none 的公司再跑限流/慢探测(Workable+Recruitee+Personio)，须绕过 none 缓存。
    # 全源提速：保留 Workable（覆盖不减），靠更高 Workable 并发 + 更多线程 + 单 slug 提速，
    # 429 交给 probe_workable 的单探测退避自节流（整体 5 分钟冷却已废除，见主循环）。
    if mode == "deep":
        probes, force, max_slugs, workers = DEEP_PROBES, True, DEEP_SLUGS, DEEP_WORKERS
    elif mode == "curated":       # 精选清单小，连 Workable 一起扫
        probes, force, max_slugs, workers = CORE_PROBES, False, SWEEP_SLUGS, SCAN_WORKERS
    else:                         # pending / all / known：只跑健康快探测
        probes, force, max_slugs, workers = PROBES, False, SWEEP_SLUGS, SCAN_WORKERS
    state["scan"] = {"running": True, "done": 0, "total": len(names), "mode": mode,
                     "started_ts": time.time(), "stop": False, "new_jobs": 0,
                     "paused_until": 0}
    _stop_event.clear()
    processed = 0

    try:
        with ThreadPoolExecutor(max_workers=workers) as pool:
            it = iter(names)
            futures = {}
            # 缓冲只比并发略多——按「停止」时要抛弃的在飞任务少，停得快
            buf = workers * 2

            def submit_more():
                while len(futures) < buf:
                    try:
                        n = next(it)
                    except StopIteration:
                        return
                    futures[pool.submit(scan_company, n, keywords, uk_only,
                                        probes, max_slugs, force)] = n

            submit_more()
            while futures:
                # 带超时轮询：深度补扫时在飞任务可能几十秒都没有一家"完成"，
                # 不加超时的话停止标志要等到有任务完成才会被检查
                done_set, _ = wait(list(futures), timeout=2,
                                   return_when=FIRST_COMPLETED)
                for fut in done_set:
                    n = futures.pop(fut)
                    try:
                        res = fut.result()
                    except Exception as e:
                        res = {"name": n, "status": "error", "error": str(e),
                               "matched": [], "scanned_at": today}
                    with _lock:
                        _record_result(n, res, today, curated)
                        state["scan"]["done"] += 1
                    processed += 1
                    if processed % 500 == 0:
                        save_ats_cache()
                        save_state()
                if state["scan"]["stop"]:
                    pool.shutdown(wait=False, cancel_futures=True)  # 取消还在排队的，秒停
                    break
                # 全源提速：不再因近期 429 触发整体 5 分钟冷却——429 由 probe_workable 的单探测
                # 退避(1.5/3/6s，4 次)就地自节流，只拖住踩雷的那几条探测，不冻结整片扫描。
                submit_more()
    finally:
        state["scan"]["running"] = False
        save_ats_cache()
        save_state()
        # 扫描结束后自动重算匹配度（有简历时），新岗位无需手动点"计算匹配度"
        if state.get("resume") and not state["match"]["running"]:
            _spawn(run_match)


def start_scan(mode):
    """mode: curated / known / pending / deep / all"""
    if state["scan"]["running"]:
        return None
    comps = state["companies"]
    curated = _curated_names()
    if mode == "curated":
        names = sorted(curated)
    elif mode == "known":   # 已探测到招聘系统的公司：快速刷新最新岗位
        names = [n for n in comps if isinstance(ats_cache.get(n), dict)]
    elif mode == "pending":  # 还没探测过的公司：全量清查（可中断续跑）
        names = [n for n in comps if n not in ats_cache]
    elif mode == "deep":    # 深度补扫：对已判 none 的公司再试慢探测(Recruitee/Personio)
        # 只挑「还没深度探过」的 none；深度探过确认没有的是 none_deep，故意排除 → 可中断续跑
        names = [n for n in comps if ats_cache.get(n) == "none"]
    else:  # all
        known = [n for n in comps if isinstance(ats_cache.get(n), dict)]
        pending = [n for n in comps if n not in ats_cache]
        names = sorted(curated) + [n for n in known if n not in curated] + pending
    if not names:
        return 0
    _charge("scan", SCAN_DAILY_CAP)   # 扣每日扫描配额（超限抛 UsageLimitError）
    _spawn(run_scan, names, mode)
    return len(names)


def auto_loop():
    """auto_hours>0 时定时快速刷新已知招聘系统，第一时间发现新岗位。
    多租户：逐个用户检查各自的 auto_hours（只覆盖当前已在内存的用户；
    某用户重启后首次登录会重新进入内存，其定时随之恢复）。"""
    while True:
        time.sleep(300)
        for uid in list(user_states.keys()):
            try:
                _ctx_user.set(uid)   # 把上下文切到这个用户，state[...] 即其数据
                h = state["auto_hours"]
                if h > 0 and not state["scan"]["running"] \
                        and time.time() - state["last_auto"] > h * 3600:
                    if start_scan("known"):   # 内部 _spawn 会捕获当前 uid
                        state["last_auto"] = time.time()
            except Exception:
                pass


# ---------------------------------------------------------------- 聚合平台搜索
# Adzuna / Reed 是关键词聚合器：几次调用就能横扫全英国所有公司/ATS 的岗位，
# 是“抓最全 + 求快”的最大杠杆。需要各自的免费官方 API key（放 data/api_keys.json）。
# 搜索词跟随用户设置的岗位关键词（改行找护理/会计也能用）；免费 API 有配额，
# 每轮最多取前 N 个关键词。
AGG_MAX_QUERIES = 6

AGG_KEY_FIELDS = ("adzuna_app_id", "reed_api_key", "jooble_api_key", "rapidapi_key")


def _has_agg_keys():
    return any(api_keys.get(k) for k in AGG_KEY_FIELDS)


def _norm_company(n):
    """公司名归一化，用于把聚合器岗位和官方 sponsor 名单交叉核对。"""
    b = re.sub(r"[^a-z0-9 ]", " ", (n or "").lower())
    words = [w for w in b.split() if w and w not in STRIP_SUFFIXES]
    return " ".join(words)


def search_adzuna(query, pages=2):
    aid, key = api_keys.get("adzuna_app_id"), api_keys.get("adzuna_app_key")
    if not aid or not key:
        return []
    out = []
    for p in range(1, pages + 1):
        url = (f"https://api.adzuna.com/v1/api/jobs/gb/search/{p}"
               f"?app_id={aid}&app_key={key}&results_per_page=50"
               f"&what={quote_plus(query)}&max_days_old=30&content-type=application/json")
        data = _get_json(url)
        res = (data or {}).get("results") or []
        for j in res:
            out.append({
                "company": (j.get("company") or {}).get("display_name", ""),
                "title": j.get("title", ""), "url": j.get("redirect_url", ""),
                "location": (j.get("location") or {}).get("display_name", ""),
                "desc": _strip_html(j.get("description", "")), "source": "Adzuna",
            })
        if len(res) < 50:
            break
    return out


def search_reed(query):
    key = api_keys.get("reed_api_key")
    if not key:
        return []
    try:
        r = _session().get("https://www.reed.co.uk/api/1.0/search",
                           params={"keywords": query, "resultsToTake": 100},
                           auth=(key, ""), timeout=TIMEOUT)
        if r.status_code != 200:
            return []
        res = r.json().get("results") or []
    except Exception:
        return []
    return [{
        "company": j.get("employerName", ""), "title": j.get("jobTitle", ""),
        "url": j.get("jobUrl", ""), "location": j.get("locationName", ""),
        "desc": _strip_html(j.get("jobDescription", "")), "source": "Reed",
    } for j in res]


def search_jooble(query):
    """Jooble 聚合了大量招聘站（含部分 LinkedIn/Indeed 转载）。免费 key，POST 调用。"""
    key = api_keys.get("jooble_api_key")
    if not key:
        return []
    try:
        r = _session().post(f"https://jooble.org/api/{key}",
                            json={"keywords": query, "location": "United Kingdom"},
                            timeout=TIMEOUT)
        if r.status_code != 200:
            return []
        res = r.json().get("jobs") or []
    except Exception:
        return []
    return [{
        "company": j.get("company", ""), "title": j.get("title", ""),
        "url": j.get("link", ""), "location": j.get("location", ""),
        "desc": _strip_html(j.get("snippet", "")), "source": "Jooble",
    } for j in res]


def search_jsearch(query):
    """JSearch(RapidAPI) 抓 Google for Jobs——覆盖 LinkedIn/Indeed/Glassdoor/ZipRecruiter，
    正是本工具官方 ATS 扫描的最大盲区。免费额度约每月 200 次。"""
    key = api_keys.get("rapidapi_key")
    if not key:
        return []
    try:
        r = _session().get("https://jsearch.p.rapidapi.com/search",
                          params={"query": f"{query} in United Kingdom",
                                  "page": "1", "num_pages": "1", "country": "gb"},
                          headers={"X-RapidAPI-Key": key,
                                   "X-RapidAPI-Host": "jsearch.p.rapidapi.com"},
                          timeout=(4, 20))
        if r.status_code != 200:
            return []
        res = r.json().get("data") or []
    except Exception:
        return []
    out = []
    for j in res:
        loc = ", ".join(x for x in (j.get("job_city"), j.get("job_country")) if x)
        out.append({
            "company": j.get("employer_name", ""), "title": j.get("job_title", ""),
            "url": j.get("job_apply_link") or j.get("job_google_link", ""),
            "location": loc, "desc": _strip_html(j.get("job_description", "")),
            "source": j.get("job_publisher") or "JSearch",
        })
    return out


def run_aggregators():
    """跑 Adzuna+Reed 关键词搜索，AI 过滤后并入 agg_jobs，并标注是否为工签 sponsor。"""
    if not _has_agg_keys():
        return
    kws = [k.lower() for k in state["keywords"]]
    queries = kws[:AGG_MAX_QUERIES]  # 跟随用户关键词，不再硬编码 AI 词
    today = date.today().isoformat()
    with _lock:
        sponsor_idx = {_norm_company(n) for n in state["companies"]}
    state["agg"] = {"running": True, "done": 0, "total": len(queries)}
    new_descs = {}  # url -> JD 摘要，存入 desc_cache 供简历匹配打分
    try:
        for q in queries:
            found = (search_adzuna(q) + search_reed(q)
                     + search_jooble(q) + search_jsearch(q))
            with _lock:
                for j in found:
                    u = j.get("url")
                    title = (j.get("title") or "").lower()
                    desc = j.get("desc") or ""
                    if not u or not match_job(title, desc, kws):
                        continue
                    if desc:
                        new_descs[u] = desc[:8000]
                    region, rlabel = classify_region(j.get("location"))
                    seen = state["agg_jobs"].get(u)
                    state["agg_jobs"][u] = {
                        "company": j.get("company", ""), "title": j.get("title", ""),
                        "url": u, "location": j.get("location", ""),
                        "source": j.get("source", ""), "region": region, "country": rlabel,
                        "level": classify_level(title),
                        "llm": kw_match(title, LLM_KEYWORDS) or kw_match(desc.lower(), LLM_KEYWORDS),
                        "sponsor": _norm_company(j.get("company", "")) in sponsor_idx,
                        "first_seen": seen["first_seen"] if seen else today,
                        "last_seen": today,
                    }
                state["agg"]["done"] += 1
    finally:
        state["agg"]["running"] = False
        if new_descs:
            cache = load_desc_cache()
            for u, d in new_descs.items():
                if len(d) > len(cache.get(u) or ""):  # 保留更完整的版本
                    cache[u] = d
            _atomic_write(DESC_CACHE_FILE, json.dumps(cache, ensure_ascii=False))
        save_state()
        # 抓完顺手把新岗位的匹配分算出来
        if state.get("resume") and not state["match"]["running"]:
            _spawn(run_match)


# ---------------------------------------------------------------- 社媒招聘帖（Hacker News）
# HN 每月有一贴 "Ask HN: Who is hiring?"（机器人 whoishiring 于每月 1 号发），
# 下面每条顶层评论就是一个招聘帖，惯例首行：Company | Role | Location | Remote | Visa | tech | url。
# Algolia 提供免费无鉴权检索接口，无需 API key，是这个工具最契合的社媒来源。
HN_ALGOLIA = "http://hn.algolia.com/api/v1"
_HN_URL_RE = re.compile(r"https?://[^\s|)<>\"']+")
_HN_VISA_RE = re.compile(r"\b(visa|sponsor|sponsorship|relocation|relocat\w*|h1b|h-1b|tier 2|skilled worker)\b", re.I)


def _hn_first_para(html_text):
    """取评论第一段（招聘帖的抬头行：公司 | 岗位 | 地点 | Remote | …）。
    HN 评论各段用 <p> 分隔，首段在第一个 <p> 之前。"""
    head = re.split(r"<p>", html_text or "", maxsplit=1)[0]
    return _strip_html(head)


def _hn_get(url):
    """HN Algolia 专用取数：宽松超时 + 一次重试。不走 ATS 那套 (2,5) 超时和限流抛错，
    因为 items 端点是 ~500KB 的大 JSON，用短超时会偶发失败被静默吞成 0。"""
    for attempt in range(2):
        try:
            r = _session().get(url, timeout=(5, 30))
            if r.status_code == 200:
                return r.json()
        except Exception:
            if attempt == 0 and _stop_event.wait(1.0):
                break
    return None


def _hn_hiring_threads(n_threads=1):
    data = _hn_get(f"{HN_ALGOLIA}/search_by_date?tags=story,author_whoishiring&hitsPerPage=8")
    hits = (data or {}).get("hits") or []
    threads = [h for h in hits if "who is hiring" in (h.get("title") or "").lower()]
    return threads[:n_threads]


def search_hackernews(n_threads=1):
    """把最近 n 期 HN 'Who is hiring' 主帖下的每条招聘评论解析成一个岗位。
    返回形状与 search_adzuna 一致：{company,title,url,location,desc,source}。"""
    out = []
    for th in _hn_hiring_threads(n_threads):
        item = _hn_get(f"{HN_ALGOLIA}/items/{th.get('objectID')}")
        for c in (item or {}).get("children") or []:
            raw = c.get("text")
            if not raw:
                continue
            plain = _strip_html(raw)
            if len(plain) < 30:       # 跳过 "[dead]"、纯 meta 之类的短评论
                continue
            header = _hn_first_para(raw)[:200] or plain[:200]
            m = _HN_URL_RE.search(plain)          # 正文里的外部申请链接
            apply_url = m.group(0) if m and "ycombinator.com" not in m.group(0) else None
            hn_url = f"https://news.ycombinator.com/item?id={c.get('id')}"
            company = re.split(r"[|:\-–—]", header)[0].strip()[:80]
            out.append({
                "company": company, "title": header,
                "url": apply_url or hn_url, "hn_url": hn_url,
                "location": header,   # 抬头行常含地点/Remote，交给 classify_region 识别
                "desc": (plain + f"  · 原帖：{hn_url}")[:8000],
                "source": "HackerNews",
            })
    return out


def run_social():
    """抓 HN 招聘帖，按用户关键词过滤后并入 agg_jobs（复用聚合器的展示/匹配管线）。
    无需任何 API key。"""
    if state["social"]["running"]:
        return
    kws = [k.lower() for k in state["keywords"]]
    today = date.today().isoformat()
    with _lock:
        sponsor_idx = {_norm_company(n) for n in state["companies"]}
    state["social"] = {"running": True, "done": 0, "total": 0, "found": 0, "err": ""}
    new_descs = {}
    added = 0
    try:
        try:
            posts = search_hackernews(n_threads=1)
        except Exception as e:
            posts = []
            state["social"]["err"] = f"抓取 HN 失败：{e}"[:200]
        state["social"]["total"] = len(posts)
        with _lock:
            for j in posts:
                u = j.get("url")
                title = (j.get("title") or "").lower()
                desc = j.get("desc") or ""
                state["social"]["done"] += 1
                if not u or not match_job(title, desc, kws):
                    continue
                added += 1
                new_descs[u] = desc[:8000]
                region, rlabel = classify_region(j.get("location"))
                seen = state["agg_jobs"].get(u)
                state["agg_jobs"][u] = {
                    "company": j.get("company", ""), "title": j.get("title", ""),
                    "url": u, "location": rlabel or "",
                    "source": "HackerNews", "region": region, "country": rlabel,
                    "level": classify_level(title),
                    "llm": kw_match(title, LLM_KEYWORDS) or kw_match(desc.lower(), LLM_KEYWORDS),
                    "sponsor": _norm_company(j.get("company", "")) in sponsor_idx,
                    "mentions_visa": bool(_HN_VISA_RE.search(desc)),
                    "first_seen": seen["first_seen"] if seen else today,
                    "last_seen": today,
                }
    finally:
        state["social"]["found"] = added
        state["social"]["running"] = False
        if new_descs:
            cache = load_desc_cache()
            for u, d in new_descs.items():
                if len(d) > len(cache.get(u) or ""):
                    cache[u] = d
            _atomic_write(DESC_CACHE_FILE, json.dumps(cache, ensure_ascii=False))
        save_state()
        if state.get("resume") and not state["match"]["running"]:
            _spawn(run_match)


# ---------------------------------------------------------------- 公司规模（大模型估）
SIZE_BANDS = ("startup", "mid", "large", "unknown")
SIZE_LABELS = {"startup": "初创 <50", "mid": "中型 50–500",
               "large": "大型 500+", "unknown": "规模未知"}
_SIZE_BATCH = 30   # 每次问大模型多少家公司


def save_company_sizes():
    _atomic_write(COMPANY_SIZE_FILE, json.dumps(company_sizes, ensure_ascii=False))


def _parse_llm_json(text):
    """从大模型输出里抠出 JSON 数组（容忍 ```json 围栏和前后废话）。"""
    t = (text or "").strip()
    if "```" in t:
        t = re.sub(r"```(?:json)?", "", t)
    m = re.search(r"\[.*\]", t, re.S)
    if not m:
        return []
    try:
        data = json.loads(m.group(0))
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _size_targets():
    """需要识别规模的公司名：有匹配岗位的 sponsor 公司 + 聚合器岗位公司，去重、跳过已缓存。"""
    names = set()
    with _lock:
        for n, r in state["results"].items():
            if r.get("matched"):
                names.add(n)
        for j in state["agg_jobs"].values():
            c = (j.get("company") or "").strip()
            if c:
                names.add(c)
    return [n for n in names if n and n not in company_sizes]


def run_company_size():
    """用大模型给"有岗位的公司"批量估员工规模档位，写入 company_sizes 并缓存。
    只问没缓存过的公司，均摊成本；不认识的模型会返回 unknown（不瞎猜）。"""
    if not _llm_ready():
        return
    todo = _size_targets()
    state["size"] = {"running": True, "done": 0, "total": len(todo)}
    try:
        for i in range(0, len(todo), _SIZE_BATCH):
            batch = todo[i:i + _SIZE_BATCH]
            prompt = (
                "给下面每家英国公司判断员工规模档位。只输出 JSON 数组，"
                '每项 {"name":..., "band":..., "note":...}。\n'
                'band 只能取："startup"(<50人), "mid"(50-500), "large"(500+), '
                '"unknown"（你不确定/没听过的必须用这个，绝不要猜）。\n'
                "note 用中文，最多6字，写判断依据或\"不确定\"。\n"
                "公司：" + json.dumps(batch, ensure_ascii=False))
            try:
                rows = _parse_llm_json(_llm_complete(prompt, max_tokens=2000))
            except Exception:
                rows = []
            by_name = {(r.get("name") or "").strip(): r for r in rows if isinstance(r, dict)}
            for n in batch:
                r = by_name.get(n) or {}
                band = r.get("band") if r.get("band") in SIZE_BANDS else "unknown"
                company_sizes[n] = {"band": band, "note": (r.get("note") or "")[:12]}
            state["size"]["done"] = min(i + len(batch), len(todo))
            save_company_sizes()
    finally:
        state["size"]["running"] = False
        save_company_sizes()


# ---------------------------------------------------------------- 简历匹配
DESC_CACHE_FILE = os.path.join(DATA_DIR, "desc_cache.json")
DESC_FAIL_FILE = os.path.join(DATA_DIR, "desc_fail.json")
FAIL_RETRY_SEC = 24 * 3600  # JD 抓取失败 24 小时后允许重试

SKILL_VOCAB = [
    # 语言
    "python", "java", "c++", "c#", "golang", "rust", "scala", "sql", "r",
    "javascript", "typescript", "matlab", "bash",
    # 机器学习
    "machine learning", "deep learning", "pytorch", "tensorflow", "keras",
    "scikit-learn", "sklearn", "xgboost", "lightgbm", "computer vision",
    "nlp", "natural language processing", "reinforcement learning",
    "recommendation", "recommender", "time series", "anomaly detection",
    "statistics", "statistical", "a/b testing", "causal inference",
    "feature engineering", "classification", "regression", "clustering",
    "neural network", "cnn", "rnn", "lstm", "gan",
    # LLM / GenAI
    "llm", "large language model", "transformer", "transformers",
    "hugging face", "huggingface", "fine-tuning", "finetuning", "lora",
    "rlhf", "prompt engineering", "rag", "retrieval augmented generation",
    "vector database", "embedding", "embeddings", "faiss", "pinecone",
    "milvus", "langchain", "llamaindex", "llama", "agent", "agentic",
    "openai", "anthropic", "claude", "gpt", "gemini", "mistral",
    "whisper", "stable diffusion", "diffusion", "multimodal",
    "speech recognition", "ocr", "generative ai", "genai", "chatbot",
    "semantic search", "knowledge graph",
    "llmops", "guardrails", "guardrail", "azure openai", "bedrock",
    "aws bedrock", "model evaluation", "llm evaluation", "model evals",
    "prompt optimization", "retrieval", "vector search", "rerank",
    "reranking", "peft", "quantization", "distillation", "vllm", "triton",
    "langgraph", "autogen", "dspy", "semantic kernel", "context window",
    "production ml", "retrieval systems",
    # 数据
    "pandas", "numpy", "spark", "pyspark", "hadoop", "kafka", "airflow",
    "dbt", "snowflake", "databricks", "etl", "data pipeline",
    "data warehouse", "bigquery", "redshift", "tableau", "power bi",
    "data analysis", "data visualization",
    # 工程 / 基础设施
    "aws", "gcp", "google cloud", "azure", "docker", "kubernetes", "k8s",
    "terraform", "ci/cd", "mlops", "mlflow", "kubeflow", "sagemaker",
    "vertex ai", "model deployment", "model serving", "inference",
    "onnx", "tensorrt", "gpu", "cuda", "distributed training", "ray",
    "monitoring", "observability", "grafana", "linux", "git", "rest api", "api",
    "bentoml", "weights & biases", "wandb", "dvc", "feature store",
    "microservices", "fastapi", "flask", "django", "react", "node.js",
    "postgresql", "mysql", "mongodb", "redis", "elasticsearch",
    "system design", "scalability", "testing", "unit test",
    # 领域 / 软技能
    "agile", "scrum", "stakeholder", "cross-functional", "mentoring",
    "leadership", "research", "publication", "phd", "kaggle",
    "open source", "fintech", "healthcare", "fraud detection", "risk",
    "personalization", "search ranking", "ads", "etl pipelines",
]
_VOCAB_RES = [(kw, re.compile(r"(?<![\w+#])" + re.escape(kw) + r"(?![\w+#])"))
              for kw in SKILL_VOCAB]

# 同义词归并：命中任一形式只记规范名
_CANON = {"sklearn": "scikit-learn", "huggingface": "hugging face",
          "finetuning": "fine-tuning", "k8s": "kubernetes",
          "natural language processing": "nlp", "google cloud": "gcp",
          "large language model": "llm", "golang": "go",
          "retrieval augmented generation": "rag",
          "generative ai": "genai", "embeddings": "embedding",
          "transformers": "transformer", "statistical": "statistics",
          "aws bedrock": "bedrock", "reranking": "rerank",
          "wandb": "weights & biases", "llm evaluation": "model evaluation",
          "model evals": "model evaluation", "guardrail": "guardrails",
          "retrieval systems": "retrieval"}


def extract_skills(text):
    t = (text or "").lower()
    found = set()
    for kw, rx in _VOCAB_RES:
        if rx.search(t):
            found.add(_CANON.get(kw, kw))
    return found


def _strip_html(html_text):
    import html as htmlmod
    txt = re.sub(r"<[^>]+>", " ", html_text or "")
    return re.sub(r"\s+", " ", htmlmod.unescape(txt)).strip()


def extract_resume_text(file_storage):
    fname = (file_storage.filename or "").lower()
    raw = file_storage.read()
    if fname.endswith(".pdf"):
        from io import BytesIO as B
        from pypdf import PdfReader
        reader = PdfReader(B(raw))
        return "\n".join((p.extract_text() or "") for p in reader.pages)
    if fname.endswith(".docx"):
        import zipfile
        from io import BytesIO as B
        with zipfile.ZipFile(B(raw)) as z:
            xml = z.read("word/document.xml").decode("utf-8", errors="replace")
        xml = re.sub(r"</w:p>", "\n", xml)
        return _strip_html(xml)
    return raw.decode("utf-8-sig", errors="replace")


def load_desc_cache():
    if os.path.exists(DESC_CACHE_FILE):
        try:
            with open(DESC_CACHE_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def load_desc_fail():
    if os.path.exists(DESC_FAIL_FILE):
        try:
            with open(DESC_FAIL_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _should_retry(u, fails):
    """抓取失败过的 URL，超过 24h 才允许再抓，避免每次匹配都重试死链。"""
    ts = fails.get(u)
    if not ts:
        return True
    try:
        return (time.time() - float(ts)) > FAIL_RETRY_SEC
    except (TypeError, ValueError):
        return True


def fetch_descriptions(r, cache, fails):
    """尽力抓取一家公司所有匹配岗位的 JD 文本，写入 cache（url -> text）。
    抓不到的记入 fails（url -> 失败时间戳），24h 后允许重试。"""
    ats, slug = r.get("ats"), r.get("slug")
    urls = {j["url"] for j in r.get("matched", []) if j.get("url")}
    todo = {u for u in urls if not cache.get(u) and _should_retry(u, fails)}
    if not todo or not ats:
        return
    try:
        if ats == "Greenhouse":
            data = _get_json(f"https://boards-api.greenhouse.io/v1/boards/{slug}/jobs?content=true")
            for j in (data or {}).get("jobs", []):
                u = j.get("absolute_url", "")
                if u in urls:
                    cache[u] = _strip_html(j.get("content", ""))[:8000]
        elif ats == "Lever":
            data = _get_json(f"https://api.lever.co/v0/postings/{slug}?mode=json")
            for j in data or []:
                u = j.get("hostedUrl", "")
                if u in urls:
                    parts = [j.get("descriptionPlain") or _strip_html(j.get("description", ""))]
                    for lst in j.get("lists", []):
                        parts.append(lst.get("text", "") + " " + _strip_html(lst.get("content", "")))
                    cache[u] = " ".join(parts)[:8000]
        elif ats == "Ashby":
            data = _get_json(f"https://api.ashbyhq.com/posting-api/job-board/{slug}?includeCompensation=false")
            for j in (data or {}).get("jobs", []):
                u = j.get("jobUrl", "") or j.get("applyUrl", "")
                if u in urls:
                    cache[u] = _strip_html(j.get("descriptionHtml", ""))[:8000]
        elif ats == "Recruitee":
            data = _get_json(f"https://{slug}.recruitee.com/api/offers/")
            for j in (data or {}).get("offers", []):
                u = j.get("careers_url", "")
                if u in urls:
                    cache[u] = _strip_html(str(j.get("description", "")) + " "
                                           + str(j.get("requirements", "")))[:8000]
        elif ats == "Workable":
            data = _get_json(f"https://apply.workable.com/api/v1/widget/accounts/{slug}")
            for j in (data or {}).get("jobs", []):
                if j.get("url") in todo and j.get("shortcode"):
                    d = _get_json(f"https://apply.workable.com/api/v1/widget/accounts/{slug}/jobs/{j['shortcode']}")
                    if d:
                        cache[j["url"]] = _strip_html(
                            str(d.get("description", "")) + " " + str(d.get("requirements", "")))[:8000]
        elif ats == "SmartRecruiters":
            for u in todo:
                jid = u.rstrip("/").rsplit("/", 1)[-1]
                d = _get_json(f"https://api.smartrecruiters.com/v1/companies/{slug}/postings/{jid}")
                if d:
                    secs = ((d.get("jobAd") or {}).get("sections") or {})
                    cache[u] = _strip_html(" ".join(
                        str(s.get("text", "")) for s in secs.values() if isinstance(s, dict)))[:8000]
    except Exception:
        pass
    for u in todo:
        if cache.get(u):
            fails.pop(u, None)          # 这次抓到了，清除失败记录
        else:
            fails[u] = time.time()      # 仍抓不到，登记失败时间，24h 后重试


# --- 贴近真实英国招聘筛选的规则 ---
NICE_SPLIT_RE = re.compile(
    r"nice to have|nice-to-have|bonus point|desirable|preferred qualification|"
    r"is a plus|would be a plus|even better if|great to have|good to have|"
    r"it would be great|we'd love|additionally")
YEARS_RES = [
    re.compile(r"(\d{1,2})\s*\+\s*years?"),
    re.compile(r"at least (\d{1,2}) years?"),
    re.compile(r"minimum (?:of )?(\d{1,2}) years?"),
    re.compile(r"(\d{1,2}) or more years?"),
]
CLEARANCE_RE = re.compile(
    r"security clearance|sc clear|dv clear|active clearance|"
    r"uk nationals? only|british citizen|must be a uk national|eligible for (?:sc|dv)\b")
NO_SPONSOR_RE = re.compile(
    r"(?:cannot|unable to|not able to|do not|don'?t|no)\s+(?:currently\s+)?"
    r"(?:offer|provide|support)?\s*(?:visa\s+)?sponsorship"
    r"|not (?:currently )?(?:able to )?sponsor|without (?:visa )?sponsorship")
YES_SPONSOR_RE = re.compile(
    r"(?:visa\s+)?sponsorship\s+(?:is\s+)?(?:available|offered|provided|supported)"
    r"|(?:we|company)\s+(?:can|do|will)\s+sponsor|able to sponsor|sponsorship for this role")


# 岗位级别 → 典型经验年限中点（用于职级匹配打分）
LEVEL_YEARS = {"grad": 0.0, "junior": 1.5, "mid": 3.5,
               "senior": 6.5, "staff": 9.0, "mgmt": 10.0}
# 领域标记：命中即代表方向对口，单独占一档权重
DOMAIN_SKILLS = {
    "nlp", "computer vision", "llm", "rag", "recommendation", "recommender",
    "fraud detection", "fintech", "healthcare", "search ranking", "ads",
    "time series", "speech recognition", "multimodal", "reinforcement learning",
    "anomaly detection", "knowledge graph", "personalization", "risk",
    "genai", "retrieval",
}

# ATS 式综合评分权重（可测的分项才计权重，最后按实际权重和归一化）：
#   必备技能 55% · 加分技能 20% · 职级 10% · 领域 10% · JD抓取置信度 5%
W_REQUIRED, W_PREFERRED, W_LEVEL, W_DOMAIN, W_CONF = 0.55, 0.20, 0.10, 0.10, 0.05


def analyze_job(title, desc, resume_kws, my_years, level=None):
    """ATS 式综合评分：技能覆盖(必备/加分) + 职级 + 领域 + JD置信度，再叠加淘汰项。

    每个分项都归一化到 0..1，只有"可测"的分项才计入权重，最后按参与权重之和
    归一化——这样缺 nice-to-have 段落或抓不到 JD 时不会被凭空扣分。"""
    dl = (desc or "").lower()
    full = (title or "").lower() + ". " + dl
    m = NICE_SPLIT_RE.search(dl)
    core_txt = dl[:m.start()] if m else dl
    nice_txt = dl[m.start():] if m else ""
    core = extract_skills((title or "").lower()) | extract_skills(core_txt)
    nice = extract_skills(nice_txt) - core
    hit_c, miss_c = core & resume_kws, core - resume_kws
    hit_n, miss_n = nice & resume_kws, nice - resume_kws

    comps = []  # (weight, subscore 0..1)
    if core:
        comps.append((W_REQUIRED, len(hit_c) / len(core)))
    if nice:
        comps.append((W_PREFERRED, len(hit_n) / len(nice)))
    if level in LEVEL_YEARS and my_years is not None:
        gap = abs(my_years - LEVEL_YEARS[level])
        comps.append((W_LEVEL, max(0.0, 1.0 - gap / 5.0)))
    job_domains = (core | nice) & DOMAIN_SKILLS
    if job_domains:
        comps.append((W_DOMAIN, len(job_domains & resume_kws) / len(job_domains)))
    # JD 抓取置信度：抓到完整正文 1.0，正文很短 0.6，只有标题 0.3
    conf = 1.0 if len(dl) >= 400 else (0.6 if dl else 0.3)
    comps.append((W_CONF, conf))

    wsum = sum(w for w, _ in comps)
    score = round(100 * sum(w * s for w, s in comps) / wsum) if wsum else 0

    flags = []
    yrs = 0
    for rx in YEARS_RES:
        for g in rx.findall(full):
            try:
                yrs = max(yrs, int(g))
            except ValueError:
                pass
    if 0 < my_years < yrs <= 15:
        flags.append({"t": f"要求 {yrs}+ 年经验", "lv": "warn"})
    if CLEARANCE_RE.search(full):
        flags.append({"t": "需安全许可/国籍要求", "lv": "warn"})
    if NO_SPONSOR_RE.search(full):
        flags.append({"t": "JD 声明不担保签证", "lv": "warn"})
        score = max(score - 40, 0)  # 真实世界里这基本等于淘汰
    elif YES_SPONSOR_RE.search(full):
        flags.append({"t": "明确提供签证担保", "lv": "good"})
    if not dl:
        # 没抓到 JD 正文：只凭标题算的分虚高（"Junior ML Engineer" 光标题就能拿 90+），
        # 封顶并明示低置信度，防止挤掉那些用完整 JD 验证过的真高分岗位
        score = min(score, 65)
        flags.append({"t": "未抓到 JD，仅按标题估分", "lv": "warn"})
    return {
        "score": score, "hit": sorted(hit_c | hit_n),
        "must_miss": sorted(miss_c)[:12], "nice_miss": sorted(miss_n)[:10],
        "flags": flags, "n": len(core) + len(nice),
        "req_years": yrs,   # JD 抽到的最低要求年限（0=未注明），供前端“经验要求”筛选
        # 分数可信度：high=完整JD / mid=JD较短 / low=只有标题——分数一样时先信 high
        "conf": ("high" if len(dl) >= 400 else "mid" if dl else "low"),
    }


def visa_certainty(ms, on_register):
    """签证确定性分层（前端一级筛选）：
      yes       —— JD 明确写提供担保
      no        —— JD 明确写不担保
      register  —— 公司在官方 sponsor 名单里，但 JD 未说明（大多数岗位在这层）
      uncertain —— 聚合平台岗位且公司名没匹配上名单，担保与否完全未知"""
    txts = [f.get("t", "") for f in ((ms or {}).get("flags") or [])]
    if any("不担保" in t for t in txts):
        return "no"
    if any("提供签证担保" in t for t in txts):
        return "yes"
    return "register" if on_register else "uncertain"


def job_tier(score, ms, visa, tmatch="target"):
    """推荐分层：strong 强推荐 / maybe 可以看看 / weak 匹配弱 /
    lowconf 低置信（没抓到JD或没算分） / risk 签证或安全许可风险。

    进 strong 必须三者同时成立：岗位方向对口(target/acceptable) + 分数≥60 +
    无签证/许可风险。tmatch=off_target 的岗位（IT Support、Sales 等）即使技能分
    很高也只归 weak——避免"技能重合"被误当成"岗位对口"。"""
    txts = [f.get("t", "") for f in ((ms or {}).get("flags") or [])]
    if visa == "no" or any("安全许可" in t for t in txts):
        return "risk"
    if tmatch == "off_target":
        return "weak"     # 方向不对：直接沉底，绝不进今日推荐（在 lowconf 判断之前）
    if score is None or any("仅按标题估分" in t for t in txts):
        return "lowconf"
    if score >= 60:
        return "strong"
    if score >= 35:
        return "maybe"
    return "weak"


# 综合推荐分权重（方向对口是精准度命门，权重高于 roadmap 原案的 0.10）
_CONF_W = {"high": 1.0, "mid": 0.6, "low": 0.3}
_VISA_W = {"yes": 1.0, "register": 0.6, "uncertain": 0.3, "no": 0.0}
_TMATCH_W = {"target": 1.0, "acceptable": 0.7, "off_target": 0.0}
TIER_ORDER = {"strong": 0, "maybe": 1, "lowconf": 2, "weak": 3, "risk": 4}


def rank_score(score, conf, visa, tmatch, days):
    """综合推荐分 0..100：技能分×0.40 + 方向×0.20 + 签证×0.20 + 置信×0.12 + 新鲜×0.08。
    tier 是粗分桶（strong/maybe/…），rank_score 用于桶内细排和"今日推荐 Top"。"""
    s = max(0, min(score or 0, 100)) / 100.0
    if days is None:
        fresh = 0.3
    elif days <= NEW_DAYS:
        fresh = 1.0
    elif days <= LONG_RUNNING_DAYS:
        fresh = 0.5
    else:
        fresh = 0.2
    val = (0.40 * s
           + 0.20 * _TMATCH_W.get(tmatch, 0.0)
           + 0.20 * _VISA_W.get(visa, 0.3)
           + 0.12 * _CONF_W.get(conf, 0.3)
           + 0.08 * fresh)
    return round(100 * val)


def run_match():
    resume = state.get("resume")
    if not resume:
        return
    resume_kws = set(resume["keywords"])
    my_years = state.get("exp_years", 1)
    cache = load_desc_cache()
    fails = load_desc_fail()
    with _lock:
        targets = [(n, dict(r)) for n, r in state["results"].items() if r.get("matched")]
        agg = [(u, dict(j)) for u, j in state["agg_jobs"].items()]
    state["match"] = {"running": True, "done": 0,
                      "total": len(targets) + (1 if agg else 0)}
    try:
        for n, r in targets:
            fetch_descriptions(r, cache, fails)
            with _lock:
                for j in r.get("matched", []):
                    u = j.get("url", "")
                    res = analyze_job(j.get("title", ""), cache.get(u, ""),
                                      resume_kws, my_years, j.get("level"))
                    if res["n"]:
                        state["match_scores"][u] = res
                state["match"]["done"] += 1
        if agg:  # 聚合平台岗位也算分：JD 用抓取时存进 desc_cache 的摘要
            agg_scores = {}
            for u, j in agg:
                res = analyze_job(j.get("title", ""), cache.get(u, ""),
                                  resume_kws, my_years, j.get("level"))
                if res["n"]:
                    agg_scores[u] = res
            with _lock:
                state["match_scores"].update(agg_scores)
                state["match"]["done"] += 1
    finally:
        state["match"]["running"] = False
        jd_ok = sum(1 for v in cache.values() if v)
        state["match_stats"] = {"jd_ok": jd_ok, "jd_fail": len(fails),
                                "updated": date.today().isoformat()}
        _atomic_write(DESC_CACHE_FILE, json.dumps(cache, ensure_ascii=False))
        _atomic_write(DESC_FAIL_FILE, json.dumps(fails, ensure_ascii=False))
        save_state()


@app.route("/api/resume", methods=["POST"])
def api_resume():
    f = request.files.get("file")
    if f:
        try:
            text = extract_resume_text(f)
        except Exception as e:
            return jsonify({"error": f"解析失败: {e}（可试试导出为 PDF/TXT 再传）"}), 400
        name = f.filename
    else:
        data = request.get_json(silent=True) or {}
        text = (data.get("text") or "").strip()
        name = "粘贴文本"
    text = (text or "").strip()
    if len(text) < 100:
        return jsonify({"error": "简历内容太短，未能识别（PDF 若是扫描件请用文字版）"}), 400
    kws = sorted(extract_skills(text))
    with _lock:
        state["resume"] = {"text": text[:20000], "keywords": kws, "name": name,
                           "updated": date.today().isoformat()}
        state["match_scores"] = {}
    save_state()
    _spawn(run_match)
    return jsonify({"ok": True, "keywords": kws, "chars": len(text)})


@app.route("/api/match", methods=["POST"])
def api_match():
    if not state.get("resume"):
        return jsonify({"error": "请先上传简历"}), 400
    if state["match"]["running"]:
        return jsonify({"error": "匹配计算进行中"}), 409
    _spawn(run_match)
    return jsonify({"started": True})


# ---------------------------------------------------------------- 简历提升建议
# 语料挖掘：抓 JD 里“experience with X / proficiency in X”之类短语，捞出词表外的技能
_FREEFORM_CUE_RE = re.compile(
    r"(?:experience (?:with|in|using)|proficien(?:t|cy) (?:in|with)|knowledge of|"
    r"familiar(?:ity)? with|expertise in|skilled? in|hands[- ]on (?:with|experience with)|"
    r"background in|understanding of|working knowledge of|deep understanding of) "
    r"([a-z0-9][a-z0-9 .+#/-]{2,40})")
_FREEFORM_STOP = {
    "the", "a", "an", "and", "or", "of", "in", "with", "using", "to", "for", "our",
    "your", "their", "this", "that", "these", "those", "such", "related", "various",
    "modern", "strong", "solid", "good", "excellent", "proven", "real", "world",
    "large", "scale", "high", "quality", "best", "practices", "practice", "years",
    "year", "team", "teams", "ability", "concepts", "principles", "at", "least",
    "is", "are", "be", "required", "preferred", "plus", "etc", "including", "as",
    "on", "per", "across", "into", "within", "other", "any", "all", "must",
    "should", "would", "will", "can", "we", "you", "they", "it", "clusters",
}


def mine_frequent_terms(descs, known, top=10):
    """从 JD 语料里挖出高频、但不在技能词表/简历里的候选技能短语（启发式）。"""
    from collections import Counter
    cnt = Counter()
    for d in descs:
        for m in _FREEFORM_CUE_RE.finditer((d or "").lower()):
            words = [w for w in re.split(r"[ /]+", m.group(1).strip())
                     if w and w not in _FREEFORM_STOP]
            for n in (2, 1):  # 优先取二元短语，退化到一元
                if len(words) >= n:
                    term = " ".join(words[:n])
                    if len(term) > 2 and term not in known:
                        cnt[term] += 1
                    break
    return cnt.most_common(top)


def resume_gap_report(top=15):
    """把简历和全平台匹配分聚合成一份缺口报告：最该补的技能 + 影响面。"""
    resume = state.get("resume")
    if not resume:
        return {"ready": False}
    from collections import Counter
    resume_kws = set(resume["keywords"])
    with _lock:
        scores = [dict(ms) for ms in state["match_scores"].values() if ms]
    weight, jobs, near, have = Counter(), Counter(), Counter(), Counter()
    n_scored = 0
    for ms in scores:
        if not ms.get("n"):
            continue
        n_scored += 1
        band = 35 <= (ms.get("score") or 0) < 70   # 中等匹配：补技能最能拉动
        for k in ms.get("must_miss", []):
            weight[k] += 2; jobs[k] += 1
            if band:
                near[k] += 1
        for k in (ms.get("nice_miss") or []):
            weight[k] += 1; jobs[k] += 1
            if band:
                near[k] += 1
        for k in ms.get("hit", []):
            have[k] += 1
    gaps = [{"skill": k, "jobs": jobs[k], "near": near[k]}
            for k, _ in weight.most_common(top)]
    known = set(SKILL_VOCAB) | set(_CANON.values()) | resume_kws
    freeform = [{"term": t, "jobs": c}
                for t, c in mine_frequent_terms(load_desc_cache().values(), known)]
    return {"ready": True, "n_scored": n_scored, "gaps": gaps,
            "have": [k for k, _ in have.most_common(15)], "freeform": freeform}


def _anthropic_key():
    return api_keys.get("anthropic_api_key") or os.environ.get("ANTHROPIC_API_KEY")


def _openai_key():
    return api_keys.get("openai_api_key") or os.environ.get("OPENAI_API_KEY")


def _llm_provider():
    """决定用哪家大模型生成建议：显式 llm_provider 优先，否则谁配了 key 用谁。"""
    pref = (api_keys.get("llm_provider") or "").strip().lower()
    if pref in ("openai", "chatgpt", "gpt"):
        return "openai" if _openai_key() else ("anthropic" if _anthropic_key() else None)
    if pref in ("anthropic", "claude"):
        return "anthropic" if _anthropic_key() else ("openai" if _openai_key() else None)
    if _openai_key():        # 未指定时：手头有 ChatGPT key 就用它，否则用 Claude
        return "openai"
    if _anthropic_key():
        return "anthropic"
    return None


def _llm_ready():
    return _llm_provider() is not None


LLM_LABELS = {"openai": "ChatGPT", "anthropic": "Claude"}
LLM_MISSING_MSG = ("未配置大模型 API key：在 data/api_keys.json 填 openai_api_key（ChatGPT）"
                   "或 anthropic_api_key（Claude），也可设环境变量 "
                   "OPENAI_API_KEY / ANTHROPIC_API_KEY")

RESUME_SYS = ("你是资深简历顾问兼 ATS 优化专家，服务在英国找 AI/ML 工作的求职者。"
              "建议要具体、可执行、诚实，绝不编造求职者没有的经历。")


def _openai_complete(prompt, max_tokens):
    from openai import OpenAI
    client = OpenAI(api_key=_openai_key())
    model = api_keys.get("openai_model", "gpt-4o")
    kw = dict(model=model,
              messages=[{"role": "system", "content": RESUME_SYS},
                        {"role": "user", "content": prompt}])
    try:
        resp = client.chat.completions.create(max_tokens=max_tokens, **kw)
    except Exception:  # 新一代/推理模型改用 max_completion_tokens
        resp = client.chat.completions.create(max_completion_tokens=max_tokens, **kw)
    return resp.choices[0].message.content or ""


def _anthropic_complete(prompt, max_tokens):
    import anthropic
    client = anthropic.Anthropic(api_key=_anthropic_key())
    model = api_keys.get("anthropic_model", "claude-opus-4-8")
    with client.messages.stream(model=model, max_tokens=max_tokens,
                                thinking={"type": "adaptive"},
                                system=RESUME_SYS,
                                messages=[{"role": "user", "content": prompt}]) as stream:
        msg = stream.get_final_message()
    return "".join(b.text for b in msg.content if getattr(b, "type", "") == "text")


def _llm_complete(prompt, max_tokens=6000):
    """按所配 provider 调用大模型生成建议，返回纯文本。
    多租户下先扣当前用户的每日 LLM 配额，超限抛 UsageLimitError。"""
    _charge("llm", LLM_DAILY_CAP)
    return (_openai_complete if _llm_provider() == "openai"
            else _anthropic_complete)(prompt, max_tokens)


def generate_resume_advice(report):
    """依据缺口报告 + 简历全文生成可执行的修改建议（Markdown）。"""
    resume = state.get("resume") or {}
    gaps_txt = "\n".join(
        f"- {g['skill']}：缺失于 {g['jobs']} 个岗位"
        + (f"（其中 {g['near']} 个是中等匹配，最可能因此被拉低）" if g["near"] else "")
        for g in report.get("gaps", []))
    free_txt = "、".join(f"{f['term']}({f['jobs']})" for f in report.get("freeform", []))
    prompt = f"""我在英国找 AI/ML 相关工作。以下是把我的简历与全平台 {report.get('n_scored', 0)} 个岗位匹配后的统计。

【我简历已覆盖的高频技能】{('、'.join(report.get('have', [])) or '（无）')}

【全平台最常要求、但我简历缺失的技能（按影响排序）】
{gaps_txt or '（无）'}

【JD 里高频出现、但可能既不在我简历也不在技能词表里的表达（启发式，仅供参考）】{free_txt or '（无）'}

【我的简历全文】
{(resume.get('text') or '')[:8000]}

请给我一份可执行的简历修改建议（中文，Markdown）：
1. 优先级：先补哪 3–5 个技能/关键词最能整体提升匹配度，并引用“缺失于多少个岗位”说明理由。
2. 每个要补的技能：若我可能已具备，建议加到简历哪一段、用什么措辞最自然（给 1 条量化 bullet 示例）。
3. 指出我简历里 2–3 条最该改写的经历，怎么改更贴近这些岗位。
4. 诚实提醒：哪些缺失技能若我确实没有，别硬写，应通过项目/课程补齐。
不要编造我没有的经历。"""
    return _llm_complete(prompt, max_tokens=8000)


@app.route("/api/resume_suggestions")
def api_resume_suggestions():
    report = resume_gap_report()
    prov = _llm_provider()
    report["has_llm"] = prov is not None
    report["llm_label"] = LLM_LABELS.get(prov)
    return jsonify(report)


@app.route("/api/resume_suggestions/generate", methods=["POST"])
def api_resume_suggestions_generate():
    if not state.get("resume"):
        return jsonify({"error": "请先上传简历"}), 400
    report = resume_gap_report()
    if not report.get("ready") or not report.get("gaps"):
        return jsonify({"error": "还没有匹配数据——请先上传简历并等匹配计算完成"}), 400
    if not _llm_ready():
        return jsonify({"error": LLM_MISSING_MSG}), 400
    try:
        advice = generate_resume_advice(report)
    except Exception as e:
        return jsonify({"error": f"生成失败: {e}"}), 500
    return jsonify({"advice": advice})


def _find_job(url):
    """按 URL 在扫描结果 / 聚合岗位里回捞岗位元信息（标题、公司、地点）。"""
    with _lock:
        for n, r in state["results"].items():
            for j in r.get("matched", []):
                if j.get("url") == url:
                    return {"title": j.get("title", ""), "company": n,
                            "location": j.get("location", "")}
        j = state["agg_jobs"].get(url)
        if j:
            return {"title": j.get("title", ""), "company": j.get("company", ""),
                    "location": j.get("location", "")}
    return None


def generate_job_advice(url):
    """针对单个岗位：拿 JD 正文 + 该岗位的匹配分析 + 简历全文，让大模型生成
    这一份岗位专属的简历修改建议（Markdown）。"""
    resume = state.get("resume") or {}
    meta = _find_job(url) or {}
    ms = state["match_scores"].get(url) or {}
    jd = (load_desc_cache().get(url) or "")[:8000]
    flags_txt = "；".join(f.get("t", "") for f in ms.get("flags", []))
    prompt = f"""请针对下面这个具体岗位，帮我优化简历，让它更容易通过 ATS 筛选并打动招聘官。

【目标岗位】{meta.get('title', '')} — {meta.get('company', '')}（{meta.get('location', '')}）
【岗位链接】{url}
【当前匹配度】{ms.get('score', '未知')}%
【JD 要求、我简历已覆盖】{('、'.join(ms.get('hit', [])) or '（无）')}
【JD 必需、我简历缺失（优先处理）】{('、'.join(ms.get('must_miss', [])) or '（无）')}
【JD 加分、我简历缺失】{('、'.join(ms.get('nice_miss', [])) or '（无）')}
{('【签证/门槛提醒】' + flags_txt) if flags_txt else ''}

【JD 正文节选】
{jd or '（未抓到正文，请只依据上面的技能缺口给建议）'}

【我的简历全文】
{(resume.get('text') or '')[:8000]}

请给出这一份岗位专属的简历修改建议（中文，Markdown）：
1. 若缺失技能我实际具备：该加到简历哪一段、用什么措辞最自然（给量化 bullet 示例）。
2. 重写我简历中最相关的 2–3 条经历，向这个岗位靠拢，量化成果、贴合 JD 关键词（利于 ATS，但别堆砌）。
3. 针对这家公司写一段 3 句话的求职信开场白。
4. 诚实提醒：哪些缺失技能若我确实没有，别硬写，建议怎么补齐。
不要编造我没有的经历。"""
    return _llm_complete(prompt, max_tokens=6000)


@app.route("/api/job_advice", methods=["POST"])
def api_job_advice():
    if not state.get("resume"):
        return jsonify({"error": "请先上传简历"}), 400
    if not _llm_ready():
        return jsonify({"error": LLM_MISSING_MSG}), 400
    data = request.get_json(silent=True) or {}
    url = (data.get("url") or "").strip()
    if not url:
        return jsonify({"error": "缺少岗位链接"}), 400
    if url not in state["match_scores"]:
        return jsonify({"error": "这个岗位还没算过匹配度，请先点「计算匹配度」"}), 400
    cached = state["advice"].get(url)
    if cached and not data.get("force"):
        return jsonify({"advice": cached, "cached": True})
    try:
        advice = generate_job_advice(url)
    except Exception as e:
        return jsonify({"error": f"生成失败: {e}"}), 500
    with _lock:
        state["advice"][url] = advice
    save_state()
    return jsonify({"advice": advice})


# ---------------------------------------------------------------- 隐藏岗位
# 挖排除词建议时跳过的通用词：这些词出现在几乎所有标题里，排除它们没有意义
_TITLE_STOP = {
    "engineer", "engineering", "developer", "scientist", "researcher", "analyst",
    "senior", "junior", "lead", "staff", "graduate", "intern", "internship",
    "the", "and", "for", "with", "remote", "hybrid", "london", "manchester",
    "data", "software", "machine", "learning", "artificial", "intelligence",
}


def exclude_suggestions():
    """从被隐藏岗位的标题里挖高频词（出现≥3次、不在关键词/排除词里），
    作为「一键加入排除词」候选——隐藏得越多，学得越准。"""
    kws = set(state["keywords"]) | set(state["exclude_keywords"])
    cnt = Counter()
    for h in state["hidden"].values():
        for w in set(re.findall(r"[a-z]{3,}", (h.get("title") or "").lower())):
            if w not in _TITLE_STOP and w not in kws:
                cnt[w] += 1
    return [w for w, c in cnt.most_common(5) if c >= 3]


@app.route("/api/hide", methods=["POST"])
def api_hide():
    data = request.get_json(silent=True) or {}
    url = (data.get("url") or "").strip()
    if not url:
        return jsonify({"error": "缺少岗位链接"}), 400
    with _lock:
        if data.get("on", True):
            state["hidden"][url] = {"ts": date.today().isoformat(),
                                    "title": (data.get("title") or "")[:200]}
        else:
            state["hidden"].pop(url, None)
        n = len(state["hidden"])
    save_state()
    return jsonify({"ok": True, "hidden": n, "excl_suggest": exclude_suggestions()})


# ---------------------------------------------------------------- 投递清单
TRACK_STATUSES = ("saved", "applied", "interview", "offer", "rejected")


@app.route("/api/track", methods=["POST"])
def api_track():
    data = request.get_json(silent=True) or {}
    url = (data.get("url") or "").strip()
    status = (data.get("status") or "").strip()
    if not url:
        return jsonify({"error": "缺少岗位链接"}), 400
    with _lock:
        if status == "remove":
            state["tracker"].pop(url, None)
        elif status in TRACK_STATUSES:
            ent = state["tracker"].get(url) or {"ts": date.today().isoformat()}
            ent["status"] = status
            # 存岗位快照：岗位下架后清单里仍能显示
            for k in ("title", "company", "location"):
                if data.get(k):
                    ent[k] = data[k]
            state["tracker"][url] = ent
        else:
            return jsonify({"error": "无效状态"}), 400
        tracker = dict(state["tracker"])
    save_state()
    return jsonify({"ok": True, "tracker": tracker})


def generate_outreach(url):
    """针对单个岗位生成个性化 networking 搭话草稿（LinkedIn 申请附言、跟进私信、冷邮件）。"""
    resume = state.get("resume") or {}
    meta = _find_job(url) or {}
    ms = state["match_scores"].get(url) or {}
    jd = (load_desc_cache().get(url) or "")[:4000]
    prompt = f"""我想就下面这个岗位，主动联系这家公司的招聘官或团队成员（networking）。请基于我的简历亮点，为我起草个性化的英文搭话内容——要像我本人手写的：具体、真诚、克制，绝不能有套话群发感。

【目标岗位】{meta.get('title', '')} — {meta.get('company', '')}（{meta.get('location', '')}）
【我与岗位的重合点】{('、'.join(ms.get('hit', [])[:10]) or '（还没算匹配度，请从我简历里挑最相关的亮点）')}
【JD 正文节选】
{jd or '（未抓到正文，依据岗位标题和公司背景来写）'}

【我的简历全文】
{(resume.get('text') or '')[:6000]}

请输出三部分（英文正文，每部分前用一行中文说明用途；Markdown 格式）：
1. **LinkedIn 好友申请附言**（≤280 字符：一句对这家公司具体产品/技术方向的真实兴趣 + 一句我是谁 + 轻量请求）
2. **通过后的跟进私信**（100–150 词：为什么对这个岗位/团队感兴趣（具体到产品或技术栈）、我最相关的 1–2 个量化成果、一个清晰的小请求——15 分钟聊聊或帮忙内推）
3. **冷邮件**（含 Subject 行；120–180 词，结构同上但更正式，结尾提一句简历见附件）

硬性要求：不编造我没有的经历；语气自信但不自夸；每一封都必须包含至少一个只属于这家公司的具体细节（产品名/技术方向/公开的近期动态），JD 里没有就基于公司名选择其公开的主营方向。"""
    return _llm_complete(prompt, max_tokens=4000)


@app.route("/api/outreach", methods=["POST"])
def api_outreach():
    if not state.get("resume"):
        return jsonify({"error": "请先上传简历"}), 400
    if not _llm_ready():
        return jsonify({"error": LLM_MISSING_MSG}), 400
    data = request.get_json(silent=True) or {}
    url = (data.get("url") or "").strip()
    if not url:
        return jsonify({"error": "缺少岗位链接"}), 400
    cached = state["outreach"].get(url)
    if cached and not data.get("force"):
        return jsonify({"draft": cached, "cached": True})
    try:
        draft = generate_outreach(url)
    except Exception as e:
        return jsonify({"error": f"生成失败: {e}"}), 500
    with _lock:
        state["outreach"][url] = draft
    save_state()
    return jsonify({"draft": draft})


# ---------------------------------------------------------------- 导入
COMPANY_HEADER_HINTS = ["company", "organisation", "organization", "employer",
                        "sponsor", "name", "公司", "雇主", "企业"]


def _pick_company_col(header):
    best = None
    for i, h in enumerate(header):
        if h is None:
            continue
        hl = str(h).lower()
        for hint in COMPANY_HEADER_HINTS:
            if hint in hl:
                if hint != "name":
                    return i
                if best is None:
                    best = i
    return best


def add_companies(names, tag, towns=None):
    added = 0
    with _lock:
        norm_existing = {n.strip().lower(): n for n in state["companies"]}
        for i, n in enumerate(names):
            n = n.strip()
            if not n:
                continue
            key = n.lower()
            if key in norm_existing:
                cur = state["companies"][norm_existing[key]]
                if tag not in cur["tags"]:
                    cur["tags"].append(tag)
            else:
                state["companies"][n] = {
                    "tags": [tag],
                    "town": (towns[i] if towns else "") or "",
                }
                norm_existing[key] = n
                added += 1
    save_companies()
    return added


@app.route("/api/import_register", methods=["POST"])
def api_import_register():
    data = request.get_json(force=True)
    path = (data.get("path") or "").strip().strip('"')
    if not os.path.exists(path):
        return jsonify({"error": f"文件不存在: {path}"}), 400
    names, towns, seen = [], [], set()
    try:
        with open(path, encoding="utf-8-sig", errors="replace") as f:
            for row in csvmod.DictReader(f):
                if (row.get("Route") or "").strip() != "Skilled Worker":
                    continue
                n = (row.get("Organisation Name") or "").strip()
                if not n:
                    continue
                key = n.lower()
                if key in seen:
                    continue
                seen.add(key)
                names.append(n)
                towns.append((row.get("Town/City") or "").strip())
    except Exception as e:
        return jsonify({"error": f"解析失败: {e}"}), 400
    added = add_companies(names, "register", towns)
    return jsonify({"parsed": len(names), "added": added,
                    "total": len(state["companies"])})


@app.route("/api/upload", methods=["POST"])
def api_upload():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "没有收到文件"}), 400
    fname = f.filename.lower()
    names = []
    try:
        if fname.endswith((".xlsx", ".xlsm")):
            import openpyxl
            wb = openpyxl.load_workbook(BytesIO(f.read()), read_only=True, data_only=True)
            for ws in wb.worksheets:
                rows = ws.iter_rows(values_only=True)
                try:
                    header = next(rows)
                except StopIteration:
                    continue
                col = _pick_company_col(header)
                if col is None:
                    continue
                for row in rows:
                    if col < len(row) and row[col]:
                        v = str(row[col]).strip()
                        if v and not v.startswith("="):
                            names.append(v)
                if names:
                    break
        else:
            text = f.read().decode("utf-8-sig", errors="replace")
            rows = list(csvmod.reader(text.splitlines()))
            if rows:
                col = _pick_company_col(rows[0])
                if col is not None and len(rows) > 1:
                    names = [r[col].strip() for r in rows[1:] if len(r) > col and r[col].strip()]
                else:
                    names = [r[0].strip() for r in rows if r and r[0].strip()]
    except Exception as e:
        return jsonify({"error": f"解析失败: {e}"}), 400
    if not names:
        return jsonify({"error": "没有识别到公司名"}), 400
    added = add_companies(names, "curated")
    return jsonify({"count": added, "total": len(state["companies"])})


@app.route("/api/companies", methods=["POST"])
def api_companies():
    data = request.get_json(force=True)
    names = [n for n in data.get("names", []) if n.strip()]
    added = add_companies(names, "curated")
    return jsonify({"added": added, "total": len(state["companies"])})


# ---------------------------------------------------------------- 路由
@app.route("/healthz")
def healthz():
    return "ok", 200


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/state")
def api_state():
    today = date.today()
    with _lock:
        curated = _curated_names()
        comps = state["companies"]
        results = state["results"]

        res_out = {}
        for n, r in results.items():
            rr = dict(r)
            rr["links"] = build_links(n)
            rr["curated"] = n in curated
            res_out[n] = rr
        # 精选公司即使没探测到招聘页也要显示
        for n in curated:
            if n not in res_out:
                st = "no_board" if ats_cache.get(n) in _NO_ATS else "pending"
                res_out[n] = {"name": n, "status": st, "matched": [],
                              "curated": True, "links": build_links(n)}

        excl = state.get("exclude_keywords") or []
        excl_re = (re.compile(r"\b(" + "|".join(re.escape(k) for k in excl) + r")\b")
                   if excl else None)
        tgt = state.get("target_titles") or DEFAULT_TARGET_TITLES
        acc = state.get("acceptable_titles") or DEFAULT_ACCEPTABLE_TITLES
        show_hidden = request.args.get("hidden") == "1"  # 已隐藏视图：只看被隐藏的

        feed = []
        for n, r in results.items():
            on_reg = "register" in (comps.get(n) or {}).get("tags", [])
            for j in r.get("matched", []):
                if (j.get("url", "") in state["hidden"]) != show_hidden:
                    continue
                if excl_re and excl_re.search((j.get("title") or "").lower()):
                    continue
                # 旧数据可能缺 region/level，即时补算
                if "region" not in j:
                    j["region"], j["country"] = classify_region(j.get("location"))
                if "level" not in j:
                    j["level"] = classify_level(j.get("title"))
                fs = j.get("first_seen", "")
                try:
                    days = (today - date.fromisoformat(fs)).days if fs else None
                except Exception:
                    days = None
                ms = state["match_scores"].get(j.get("url", ""))
                visa = visa_certainty(ms, on_reg)
                tmatch = title_match(j.get("title", ""), tgt, acc)
                feed.append({
                    "title": j.get("title", ""), "url": j.get("url", ""),
                    "location": j.get("location", ""), "company": n,
                    "curated": n in curated, "llm": bool(j.get("llm")),
                    "region": j["region"], "country": j.get("country", ""),
                    "level": j["level"], "via": j.get("via", "title"),
                    "first_seen": fs, "days": days,
                    "njobs": r.get("total_jobs", 0),
                    "csize": (company_sizes.get(n) or {}).get("band", ""),
                    "req_years": ms.get("req_years", 0) if ms else 0,
                    "score": ms["score"] if ms else None,
                    "conf": ms.get("conf") if ms else None,
                    "visa": visa, "title_match": tmatch,
                    "tier": job_tier(ms["score"] if ms else None, ms, visa, tmatch),
                    "rank": rank_score(ms["score"] if ms else None,
                                       ms.get("conf") if ms else None, visa, tmatch, days),
                    "hit": ms["hit"] if ms else [],
                    "must_miss": (ms.get("must_miss") or ms.get("miss", [])) if ms else [],
                    "nice_miss": ms.get("nice_miss", []) if ms else [],
                    "flags": ms.get("flags", []) if ms else [],
                })
        # 并入聚合平台(Adzuna/Reed)岗位
        for u, j in state["agg_jobs"].items():
            if (u in state["hidden"]) != show_hidden:
                continue
            if excl_re and excl_re.search((j.get("title") or "").lower()):
                continue
            region = j.get("region", "uk")
            if state["uk_only"] and region == "other":
                continue
            fs = j.get("first_seen", "")
            try:
                days = (today - date.fromisoformat(fs)).days if fs else None
            except Exception:
                days = None
            ms = state["match_scores"].get(u)
            visa = visa_certainty(ms, bool(j.get("sponsor")))
            tmatch = title_match(j.get("title", ""), tgt, acc)
            feed.append({
                "title": j.get("title", ""), "url": u,
                "location": j.get("location", ""), "company": j.get("company", ""),
                "curated": False, "llm": bool(j.get("llm")),
                "region": region, "country": j.get("country", ""),
                "level": j.get("level", "mid"), "via": "agg",
                "source": j.get("source", ""), "sponsor": bool(j.get("sponsor")),
                "first_seen": fs, "days": days,
                "njobs": None,
                "csize": (company_sizes.get(j.get("company", "")) or {}).get("band", ""),
                "req_years": ms.get("req_years", 0) if ms else 0,
                "score": ms["score"] if ms else None,
                "conf": ms.get("conf") if ms else None,
                "visa": visa, "title_match": tmatch,
                "tier": job_tier(ms["score"] if ms else None, ms, visa, tmatch),
                "rank": rank_score(ms["score"] if ms else None,
                                   ms.get("conf") if ms else None, visa, tmatch, days),
                "hit": ms["hit"] if ms else [],
                "must_miss": ms.get("must_miss", []) if ms else [],
                "nice_miss": ms.get("nice_miss", []) if ms else [],
                "flags": ms.get("flags", []) if ms else [],
            })
        # 按前端请求的排序方式排好再截断，避免"高分老岗位"被时间序截断挡在 400 名外
        sort = request.args.get("sort")
        if sort == "rank":   # 推荐优先：先按 tier 分桶优先级，桶内按综合分降序
            feed.sort(key=lambda x: (TIER_ORDER.get(x["tier"], 9), -(x.get("rank") or 0),
                                     x["first_seen"] or ""))
        elif sort == "score":
            feed.sort(key=lambda x: (x["score"] if x["score"] is not None else -1,
                                     x["first_seen"] or ""), reverse=True)
        else:
            feed.sort(key=lambda x: (x["first_seen"] or ""), reverse=True)

        probed = len(ats_cache)
        n_reg = sum(1 for c in comps.values() if "register" in c["tags"])
        mstat = state.get("match_stats") or {}
        jd_ok, jd_fail = mstat.get("jd_ok", 0), mstat.get("jd_fail", 0)
        counts = {
            "total": len(comps), "curated": len(curated), "register": n_reg,
            "probed": probed,
            "pending": sum(1 for n in comps if n not in ats_cache),
            "with_board": len(results),
            "slug_failed": sum(1 for v in ats_cache.values() if v in _NO_ATS),
            "overrides": len(slug_overrides),
            "has_ai": sum(1 for r in results.values() if r["status"] == "has_ai"),
            "ai_jobs": len(feed),
            "llm_jobs": sum(1 for j in feed if j["llm"]),
            "via_jd_jobs": sum(1 for j in feed if j.get("via") == "jd"),
            "agg_jobs": sum(1 for j in feed if j.get("via") == "agg"),
            "agg_sponsor": sum(1 for j in feed if j.get("via") == "agg" and j.get("sponsor")),
            "uk_jobs": sum(1 for j in feed if j["region"] == "uk"),
            "eu_jobs": sum(1 for j in feed if j["region"] == "europe"),
            "new_jobs": sum(1 for j in feed if j["days"] is not None and j["days"] <= NEW_DAYS),
            "long_jobs": sum(1 for j in feed if j["days"] is not None and j["days"] >= LONG_RUNNING_DAYS),
            "jd_ok": jd_ok, "jd_fail": jd_fail,
            "jd_rate": (round(100 * jd_ok / (jd_ok + jd_fail)) if (jd_ok + jd_fail) else None),
            "hidden_jobs": len(state["hidden"]),
        }
        scan = dict(state["scan"])
        if scan["running"] and scan["done"]:
            rate = scan["done"] / max(time.time() - scan["started_ts"], 1)
            scan["eta_min"] = round((scan["total"] - scan["done"]) / max(rate, 0.01) / 60)
        try:
            limit = max(50, min(int(request.args.get("limit", 400)), 5000))
        except ValueError:
            limit = 400
        resume = state.get("resume")
        return jsonify({
            "keywords": state["keywords"], "uk_only": state["uk_only"],
            "exclude_keywords": state["exclude_keywords"],
            "target_titles": tgt, "acceptable_titles": acc,
            "excl_suggest": exclude_suggestions(),
            "auto_hours": state["auto_hours"], "exp_years": state["exp_years"],
            "counts": counts, "results": res_out, "feed": feed[:limit],
            "scan": scan, "match": dict(state["match"]), "agg": dict(state["agg"]),
            "size": dict(state["size"]), "social": dict(state["social"]),
            "has_api_keys": _has_agg_keys(),
            "has_llm": _llm_ready(),
            "llm_label": LLM_LABELS.get(_llm_provider()),
            "tracker": dict(state["tracker"]),
            "resume": ({"name": resume["name"], "updated": resume["updated"],
                        "keywords": resume["keywords"], "text": resume["text"]}
                       if resume else None),
            "consts": {"new_days": NEW_DAYS, "long_days": LONG_RUNNING_DAYS,
                       "level_labels": LEVEL_LABELS, "size_labels": SIZE_LABELS},
            "auth": _auth_info(),
        })


def _auth_info():
    """给前端：当前登录名、是否多用户、以及当日剩余配额（用于显示退出/额度）。"""
    uid = _uid()
    info = {"multiuser": MULTIUSER, "user": (uid if MULTIUSER else None)}
    if MULTIUSER and not (OWNER and uid == OWNER):
        with _usage_lock:
            u = _usage_today(uid)
            info["usage"] = {"llm": u["llm"], "llm_cap": LLM_DAILY_CAP,
                             "scan": u["scan"], "scan_cap": SCAN_DAILY_CAP}
    return info


@app.route("/api/settings", methods=["POST"])
def api_settings():
    data = request.get_json(force=True)
    with _lock:
        if "keywords" in data:
            kws = [k.strip().lower() for k in data["keywords"] if k.strip()]
            if kws:
                state["keywords"] = kws
        if "exclude_keywords" in data:  # 传空列表 = 清空排除词
            state["exclude_keywords"] = [
                k.strip().lower() for k in data["exclude_keywords"] if k.strip()]
        if "target_titles" in data:      # 传空列表 = 恢复默认目标标题
            state["target_titles"] = [
                k.strip().lower() for k in data["target_titles"] if k.strip()]
        if "acceptable_titles" in data:
            state["acceptable_titles"] = [
                k.strip().lower() for k in data["acceptable_titles"] if k.strip()]
        if "uk_only" in data:
            state["uk_only"] = bool(data["uk_only"])
        if "auto_hours" in data:
            try:
                state["auto_hours"] = max(0, int(data["auto_hours"]))
            except Exception:
                pass
        if "exp_years" in data:
            try:
                state["exp_years"] = max(0, int(data["exp_years"]))
            except Exception:
                pass
    save_state()
    return jsonify({"ok": True})


@app.route("/api/scan", methods=["POST"])
def api_scan():
    data = request.get_json(silent=True) or {}
    mode = data.get("mode", "known")
    n = start_scan(mode)
    if n is None:
        return jsonify({"error": "已有扫描在进行中"}), 409
    if n == 0:
        return jsonify({"error": "该模式下没有待扫描的公司"}), 400
    return jsonify({"started": True, "total": n, "mode": mode})


@app.route("/api/scan/stop", methods=["POST"])
def api_scan_stop():
    state["scan"]["stop"] = True
    _stop_event.set()  # 打断在飞任务的信号量排队/退避 sleep，让扫描秒级收尾
    return jsonify({"ok": True})


@app.route("/api/slug_override", methods=["POST"])
def api_slug_override():
    """手动修正某公司的 ATS 定位（slug 猜错、或 Workday 这类必须手配的）。
    body: {name, ats, slug}  或  {name, ats:"Workday", host, site}。
    slug/host 都传空 → 删除该 override。"""
    data = request.get_json(force=True)
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "缺少公司名"}), 400
    key = name.lower()
    ats = (data.get("ats") or "").strip()
    slug = (data.get("slug") or "").strip()
    host = (data.get("host") or "").strip()
    site = (data.get("site") or "").strip()
    if not slug and not (host and site):
        slug_overrides.pop(key, None)
        ats_cache.pop(name, None)          # 清掉旧缓存，下次重新探测
    else:
        if ats not in OVERRIDE_ATS:
            return jsonify({"error": f"ats 需为 {sorted(OVERRIDE_ATS)} 之一"}), 400
        if ats == "Workday":
            if not (host and site):
                return jsonify({"error": "Workday 需要 host 和 site"}), 400
            slug_overrides[key] = {"ats": "Workday", "host": host, "site": site}
        else:
            if not slug:
                return jsonify({"error": f"{ats} 需要 slug"}), 400
            slug_overrides[key] = {"ats": ats, "slug": slug}
        ats_cache.pop(name, None)          # 让下次扫描按 override 走
    save_slug_overrides()
    save_ats_cache()
    return jsonify({"ok": True, "overrides": len(slug_overrides)})


@app.route("/api/aggregators", methods=["POST"])
def api_aggregators():
    """触发 Adzuna/Reed 聚合搜索，按用户关键词横扫全英国岗位。"""
    if not _has_agg_keys():
        return jsonify({"error": "未配置 Adzuna/Reed 的 API key（见 data/api_keys.json）"}), 400
    if state["agg"]["running"]:
        return jsonify({"error": "聚合搜索进行中"}), 409
    _spawn(run_aggregators)
    return jsonify({"started": True,
                    "queries": len(state["keywords"][:AGG_MAX_QUERIES])})


@app.route("/api/social", methods=["POST"])
def api_social():
    """抓 Hacker News 'Who is hiring' 招聘帖，按关键词过滤后并入岗位流。无需 API key。"""
    if state["social"]["running"]:
        return jsonify({"error": "社媒招聘帖抓取进行中"}), 409
    _spawn(run_social)
    return jsonify({"started": True})


@app.route("/api/company_size", methods=["POST"])
def api_company_size():
    """用大模型给有岗位的公司批量估员工规模，结果缓存。"""
    if not _llm_ready():
        return jsonify({"error": LLM_MISSING_MSG}), 400
    if state["size"]["running"]:
        return jsonify({"error": "公司规模识别进行中"}), 409
    n = len(_size_targets())
    if not n:
        return jsonify({"error": "没有需要识别的新公司（有岗位的都已识别过）"}), 400
    _spawn(run_company_size)
    return jsonify({"started": True, "total": n})


@app.route("/api/clear", methods=["POST"])
def api_clear():
    # 多租户：只清「我」的扫描数据；公司总目录是共享的，不在这里动
    # （否则一个人清空会抹掉所有人共享的 12 万公司目录）。
    with _lock:
        state["results"] = {}
        state["jobs_seen"] = {}
        state["agg_jobs"] = {}
        state["match_scores"] = {}
    save_state()
    return jsonify({"ok": True})


@app.route("/api/export")
def api_export():
    import openpyxl
    from openpyxl.styles import Font
    today = date.today()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI Jobs"
    head = ["Company", "精选", "Job Title", "级别", "LLM", "地区", "Location",
            "首次发现", "在线天数", "长期在招", "Job URL", "ATS", "Job Board",
            "LinkedIn People Search", "Hunter.io"]
    ws.append(head)
    for c in ws[1]:
        c.font = Font(bold=True)
    with _lock:
        results = dict(state["results"])
        curated = _curated_names()
    rows = []
    for name, r in results.items():
        links = build_links(name)
        for j in r.get("matched", []):
            fs = j.get("first_seen", "")
            try:
                days = (today - date.fromisoformat(fs)).days if fs else ""
            except Exception:
                days = ""
            level = j.get("level") or classify_level(j.get("title"))
            region = j.get("region")
            country = j.get("country", "")
            if not region:
                region, country = classify_region(j.get("location"))
            rows.append([
                name, "⭐" if name in curated else "", j.get("title", ""),
                LEVEL_LABELS.get(level, level), "⭐" if j.get("llm") else "",
                country or ("UK" if region in ("uk", "unknown") else region),
                j.get("location", ""), fs, days,
                "长期" if isinstance(days, int) and days >= LONG_RUNNING_DAYS else "",
                j.get("url", ""), r.get("ats") or "", r.get("board_url") or "",
                links["li_people"], links["hunter"],
            ])
    rows.sort(key=lambda x: (x[7] or ""), reverse=True)
    for row in rows:
        ws.append(row)
    for col, w in zip("ABCDEFGHIJKLMNO", [28, 6, 45, 10, 6, 12, 25, 12, 9, 9, 60, 14, 40, 60, 40]):
        ws.column_dimensions[col].width = w
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"AI_Job_Scout_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    load_users()  # 载入多账号表（data/users.json + 兼容老环境变量）
    load_all()
    seed_builtin_companies()  # 预置大厂/AI 雇主（curated），首次启动即可扫到
    if state["companies"] and not os.path.exists(COMPANIES_FILE):
        save_companies()  # 从旧格式迁移时立即落盘，防止丢名单
    threading.Thread(target=auto_loop, daemon=True).start()
    port = int(os.environ.get("PORT", "5050"))
    # 默认只绑本机（安全）；容器/服务器里设 HOST=0.0.0.0 才能对外访问。
    host = os.environ.get("HOST", "127.0.0.1")
    print(f"\n  AI Job Scout running:  http://127.0.0.1:{port}")
    print("  This window is the website's server — keep it open while you use the tool (you can minimize it).\n")
    if FROZEN:  # exe 双击启动时自动打开浏览器
        threading.Timer(1.5, lambda: webbrowser.open(f"http://127.0.0.1:{port}")).start()
    try:
        app.run(host=host, port=port, debug=False, threaded=True)
    except OSError:
        # 端口被占用：多半是工具已经开着了，直接打开页面即可
        print("  看起来工具已经在运行了，直接打开浏览器页面。")
        webbrowser.open(f"http://127.0.0.1:{port}")
        time.sleep(3)
